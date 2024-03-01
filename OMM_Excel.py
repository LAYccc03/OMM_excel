from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Color
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl import load_workbook
from PIL import Image as PILImage

import os
import re
import sys
import time
import json
import math
import shutil
import difflib
import hashlib
import chardet
import requests
import tempfile
import subprocess
from datetime import datetime
from PyQt5.QtWidgets import QGraphicsDropShadowEffect
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QLineEdit, QFileDialog, QVBoxLayout, QApplication, QLabel, QMessageBox, QDialog, QGridLayout
from PyQt5.QtWidgets import (QTimeEdit,QHBoxLayout,QListWidget,QCheckBox,QDialogButtonBox,QListWidgetItem,QAction,QMenu,QComboBox,QSpacerItem,QSizePolicy,QProgressDialog,QTextEdit)
from PyQt5.QtCore import QThread, pyqtSignal, QTimer, Qt, QSize, QTime,QDateTime
from PyQt5.QtGui import QFont, QColor,QIcon,QPixmap,QCursor,QImage,QValidator,QPainter,QLinearGradient
from openpyxl import load_workbook
from random import uniform
from hashlib import md5
from copy import copy

def resource_path(relative_path):
    """ 获取资源的绝对路径，用于打包后的环境 """
    if hasattr(sys, '_MEIPASS'):
        # 如果是在打包环境中
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

current_version = "4.3" #当前程序版本号

data_file_version = "3.0" #推荐使用的Excel模板

def find_rows(ws):
    # 初始化行变量为None
    row_dimension, row_instrument, row_nominal, row_usl, row_lsl, row_parts_name, row_inspection_date, row_inspection_machine = None, None, None, None, None, None, None, None
    row_diagram = None  # 用于存储“测量示意图”所在行
    # 遍历工作表中的每一行
    for row in ws.iter_rows():
        # 遍历行中的每一个单元格
        for cell in row:
            # 检查单元格是否包含字符串类型的值
            if cell.value and isinstance(cell.value, str):
                # 根据单元格的值，确定行的类型
                if "Dimension" in cell.value:
                    row_dimension = cell.row  # 记录尺寸行的行号
                elif "Instrument" in cell.value:
                    row_instrument = cell.row + 1  # 记录仪器行的下一行的行号
                elif "Nominal" in cell.value:
                    row_nominal = cell.row  # 记录标称行的行号
                elif "USL" in cell.value:
                    row_usl = cell.row  # 记录USL(上限规格)行的行号
                elif "LSL" in cell.value:
                    row_lsl = cell.row  # 记录LSL(下限规格)行的行号
                elif "Parts Name" in cell.value:
                    row_parts_name = cell.row  # 记录零件名称行的行号
                elif "Inspection Date" in cell.value:
                    row_inspection_date = cell.row  # 记录检查日期行的行号
                elif "测量示意图" in cell.value:  # 检查“测量示意图”文本
                    row_diagram = cell.row
                elif "检验机台" in cell.value:  # 检查“检验机台”文本
                    row_inspection_machine = cell.row  # 记录“检验机台”行的行号
        
        # 如果所有类型的行都找到了，就中断循环
        if all([row_dimension, row_instrument, row_nominal, row_usl, row_lsl, row_parts_name, row_inspection_date, row_diagram, row_inspection_machine]):
            break
    
    # 返回找到的行号，包括“检验机台”行
    return row_dimension, row_instrument, row_nominal, row_usl, row_lsl, row_parts_name, row_inspection_date, row_diagram, row_inspection_machine



def find_name_cell(ws):
    cell_info = {
        'type': None, 
        'name': None,
        'standard_value': None,
        'upper_tolerance': None,
        'lower_tolerance': None,
        'measured_value': None,
        'error': None,
        'judgement': None
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value == "类型":
                    cell_info['type'] = (cell.row, cell.column)
                elif cell.value == "名称":
                    cell_info['name'] = (cell.row, cell.column)
                elif cell.value == "标准值":
                    cell_info['standard_value'] = (cell.row, cell.column)
                elif cell.value == "正公差":
                    cell_info['upper_tolerance'] = (cell.row, cell.column)
                elif cell.value == "负公差":
                    cell_info['lower_tolerance'] = (cell.row, cell.column)
                elif cell.value == "测量值":
                    cell_info['measured_value'] = (cell.row, cell.column)
                elif cell.value == "误差":
                    cell_info['error'] = (cell.row, cell.column)
                elif cell.value == "判定":
                    cell_info['judgement'] = (cell.row, cell.column)
    
    return cell_info


def find_third_sheet_cells(ws):
    cell_info_3rd = {
        'name': None, 
        'type': None,
        'diameter': None,
        'x_standard_value': None,
        'y_standard_value': None,
        'x_measured_value': None,
        'y_measured_value': None,
        'x_error': None,
        'y_error': None,
        'concentricity': None,
        'concentricity_tolerance': None  # 添加“同心度公差”列的获取
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value == "名称":
                    cell_info_3rd['name'] = (cell.row, cell.column)
                elif cell.value == "类型":
                    cell_info_3rd['type'] = (cell.row, cell.column)
                elif cell.value == "直径":
                    cell_info_3rd['diameter'] = (cell.row, cell.column)
                elif cell.value == "X 标准值":
                    cell_info_3rd['x_standard_value'] = (cell.row, cell.column)
                elif cell.value == "Y 标准值":
                    cell_info_3rd['y_standard_value'] = (cell.row, cell.column)
                elif cell.value == "X 测量值":
                    cell_info_3rd['x_measured_value'] = (cell.row, cell.column)
                elif cell.value == "Y 测量值":
                    cell_info_3rd['y_measured_value'] = (cell.row, cell.column)
                elif cell.value == "X 误差":
                    cell_info_3rd['x_error'] = (cell.row, cell.column)
                elif cell.value == "Y 误差":
                    cell_info_3rd['y_error'] = (cell.row, cell.column)
                elif cell.value == "同心度":
                    cell_info_3rd['concentricity'] = (cell.row, cell.column)
                elif cell.value == "同心度公差":  # 检测并获取“同心度公差”列的位置
                    cell_info_3rd['concentricity_tolerance'] = (cell.row, cell.column)
    
    return cell_info_3rd

def find_name_cell_page4(ws):
    cell_info_page4 = {
        'type': None,
        'diameter': None,
        'standard_value': None,
        'measured_value': None
    }

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value == "类型":
                    cell_info_page4['type'] = (cell.row, cell.column)
                elif cell.value == "直径":
                    cell_info_page4['diameter'] = (cell.row, cell.column)
                elif cell.value == "标准值":
                    cell_info_page4['standard_value'] = (cell.row, cell.column)
                elif cell.value == "测量值":
                    cell_info_page4['measured_value'] = (cell.row, cell.column)
    
    return cell_info_page4



def read_txt_lines(txt_file_path):
    with open(txt_file_path, 'rb') as file:
        raw_data = file.read()
    detected_encoding = chardet.detect(raw_data)['encoding']
    
    # 使用检测到的编码重新读取文件
    with open(txt_file_path, 'r', encoding=detected_encoding) as file:
        lines = [line.strip() for line in file.readlines() if line.strip()]
    return lines


def process_lines(lines):
    projects = []  # 用于存储所有项目
    current_project = []  # 用于存储当前处理的项目

    for line in lines:
        if line == ':BEGIN':
            current_project = []  # 开始新项目时重置当前项目
        elif line == ':END':
            projects.append(current_project)  # 结束当前项目时，将其添加到项目列表
        else:
            # 将行分割为制表符分隔的元素，并添加到当前项目
            current_project.append(line.split('\t'))
    return projects


def process_project(project, exclude_center_coordinates=False):
    filtered_columns = []  # 用于存储处理后的列数据
    prompt_data = {}  # 用于存储提示信息

    for line in project:
        col_0 = line[0].split('"')[1]  # 分割第一列并取第二部分
        # 对剩余列进行处理，空或None的值直接跳过
        cols_rest = line[1:]

        # 如果第一列以"提示"开头，则处理提示信息
        if col_0.startswith("提示"):
            keyword = col_0.split(": ")[1]
            prompt_data[keyword] = cols_rest[0].replace('"', '')  # 存储提示信息
        else:
            # 检查cols_rest的第一个和第二个元素是否为空或None，如果是则跳过
            if not cols_rest[0] or not cols_rest[1]:
                continue
            # 如果设定了排除“中心坐标”，并且当前列包含“中心坐标”
            if exclude_center_coordinates and "中心坐标" in col_0:
                continue
            # 将处理后的列数据添加到列表
            filtered_columns.append((col_0, *cols_rest))

    return filtered_columns, prompt_data



# 定义用于清理文件名的函数
def clean_filename(filename):
    # 定义一个安全的字符集，包括字母、数字、下划线、短横线、括号和空格
    safe_chars = re.compile(r'[^a-zA-Z0-9_\- \(\)\（\）\.\·\+\#\u4e00-\u9fff-]')
    # 移除不在安全字符集中的所有字符
    return safe_chars.sub('', filename)

def update_formula_column(formula, old_col, new_col):
    """更新公式中的列引用"""
    def replace_col(match):
        # 替换列部分，但保留行号不变
        return new_col + match.group(2)
    
    col_regex = f"({old_col})(\d+)"
    return re.sub(col_regex, replace_col, formula)


def parse_range_limit(limit_str):
    if limit_str.startswith("<"):
        symbol = "<="
        value = limit_str[1:]
    elif limit_str.startswith(">"):
        symbol = ">="
        value = limit_str[1:]
    elif limit_str.startswith("="):
        symbol = "="
        value = limit_str[1:]
    else:
        symbol = "="
        value = limit_str

    try:
        value = float(value)
        return symbol, value
    except ValueError:
        return None, None  # 返回None表示无法解析


def write_to_excel(ws, projects, rows, filename_without_extension, judgement_type, selected_method, range_limit, machine_number, consider_xy_center, second_sheet_index, cell_info, third_sheet_index, cell_info_3rd, fourth_sheet_index, cell_info_4th):
    row_dimension, row_instrument, row_nominal, row_usl, row_lsl, row_parts_name, row_inspection_date, row_diagram, row_inspection_machine = rows

    print(judgement_type, selected_method, range_limit)

    # 先将judgement_type中的“偏移量”替换回“位置度”
    judgement_type = judgement_type.replace("偏移量", "位置度")


    # 其他样式和格式设置
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font_dimension = Font(size=8)
    font_39 = Font(size=9)
    font_nominal = Font(name='Arial', size=9, bold=True)
    fill_2 = PatternFill(fill_type='solid', fgColor='CCFFCC')

    col_start = 9  # 更新开始列索引

    earliest_date = None

    latest_date = None  # 用于存储最晚的日期

    # 检查judgement_type是否为空或包含'/'
    if judgement_type and '/' in judgement_type:
        primary_type, linked_type = judgement_type.split('/')
    else:
        primary_type = judgement_type if judgement_type else ""
        linked_type = ""

    # 检查是否有可用的range_limit
    if range_limit:
        # 将range_limit分割并尝试转换为浮点数列表，仅保留前六个值
        range_limits = []
        for rl in range_limit.split('/')[:6]:
            symbol, value = parse_range_limit(rl)
            if symbol is not None and value is not None:
                range_limits.append((symbol, value))
    else:
        range_limits = []

    # 初始化存储每个限制条件下所有项目的最大值和对应名称的字典
    all_max_values = {}
    all_max_names = {}
    all_max_usl_values = {}
    all_max_lsl_values = {}
    all_max_nominal_values = {}

    for project_idx, project in enumerate(projects):
        # 使用consider_xy_center状态决定是否排除“中心坐标”
        filtered_columns, prompt_data = process_project(project, exclude_center_coordinates=not consider_xy_center)

        # Handle Prompt Data (输入 and 日期/时间)
        if "输入" in prompt_data:
            cell = ws.cell(row=row_instrument + project_idx, column=1, value=prompt_data["输入"])
            cell.font = Font(size=12, bold=True)

        if "日期/时间" in prompt_data:
            cell = ws.cell(row=row_instrument + project_idx, column=2, value=prompt_data["日期/时间"])
            cell.font = Font(size=9, name='Arial')
            # Adjust column width for this specific cell to fit the content
            ws.column_dimensions[get_column_letter(cell.column)].auto_size = True
            ws.column_dimensions[get_column_letter(cell.column)].width = len(prompt_data["日期/时间"]) + 2  # Adding some extra space

            # 解析日期时间
            date_time_str = prompt_data["日期/时间"]
            try:
                date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
                # 检查并更新最早的日期时间
                if earliest_date is None or date_time_obj < earliest_date:
                    earliest_date = date_time_obj

                # 检查并更新最晚的日期时间
                if latest_date is None or date_time_obj > latest_date:
                    latest_date = date_time_obj
            except ValueError:
                # 如果日期时间格式不正确，这里可以添加错误处理
                pass

        inspection_machine_cell = ws.cell(row=row_inspection_machine, column=2, value=f"{machine_number}#")
        inspection_machine_cell.font = Font(size=12, bold=True)


        if '/' in judgement_type:
            primary_type, linked_type = judgement_type.split('/')
            for i, (symbol, limit) in enumerate(range_limits):
                # 如果linked_type包含“位置度”，则替换为“偏移量”
                display_linked_type = linked_type.replace("位置度", "偏移量")
                combined_value = f"{primary_type}({symbol} {limit})\n{display_linked_type + selected_method}"
                user_type_cell = ws.cell(row=row_instrument - 1, column=3 + i, value=combined_value)
                user_type_cell.alignment = center_alignment
                user_type_cell.font = Font(name='Arial', size=10, bold=True)
        else:
            # 如果judgement_type包含“位置度”，则替换为“偏移量”
            display_judgement_type = judgement_type.replace("位置度", "偏移量")
            if range_limits:
                symbol, limit = range_limits[0]
                combined_value = f"{display_judgement_type + selected_method}({symbol} {limit})"
                user_type_cell = ws.cell(row=row_instrument - 1, column=3, value=combined_value)
                user_type_cell.alignment = center_alignment
                user_type_cell.font = Font(name='Arial', size=10, bold=True)
            else:
                return  # 没有range_limits，退出函数

        # 取消合并
        if ws.merged_cells.ranges:
            for merged_cell in ws.merged_cells.ranges:
                if merged_cell.min_row == row_parts_name and merged_cell.min_col == 3:
                    ws.unmerge_cells(str(merged_cell))
                    break

        # 更新单元格的值
        cell = ws.cell(row=row_parts_name, column=3, value=filename_without_extension)
        cell.font = Font(name='Arial', size=10, bold=True, color=Color(rgb="FFFFFF"))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
    
        # 重新合并
        ws.merge_cells(start_row=row_parts_name, start_column=3, end_row=row_parts_name, end_column=5)

        # 取消"Inspection Date"旁边单元格的合并
        if ws.merged_cells.ranges:
            for merged_cell in ws.merged_cells.ranges:
                if merged_cell.min_row == row_inspection_date and merged_cell.min_col == 3:
                    ws.unmerge_cells(str(merged_cell))
                    break

        # 如果存在最晚日期，则使用该日期，否则使用当前日期
        date_to_write = earliest_date.strftime('%Y-%m-%d') if earliest_date else datetime.now().strftime('%Y-%m-%d')

        # 写入最晚的日期（格式为 %Y-%m-%d）
        cell = ws.cell(row=row_inspection_date, column=3, value=date_to_write)
        cell.font = Font(name='Arial', size=12, bold=True, color=Color(rgb="FFFFFF"))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 重新合并单元格
        ws.merge_cells(start_row=row_inspection_date, start_column=3, end_row=row_inspection_date, end_column=5)

        max_values_per_project = {}
        max_names_per_project = {}

        # 分解 judgement_type
        if '/' in judgement_type:
            primary_type, linked_type = judgement_type.split('/')
        else:
            primary_type, linked_type = judgement_type, ""


        # 遍历所有的限制信息
        for limit_index, limit_info in enumerate(range_limits):
            operator, limit = limit_info[0], limit_info[1] if len(limit_info) > 1 else limit_info[0]
            filtered_ids = []

            # 准备前一个和下一个限制值的值（如果存在）
            prev_limit_value = range_limits[limit_index - 1][1] if limit_index > 0 else None
            next_limit_value = range_limits[limit_index + 1][1] if limit_index + 1 < len(range_limits) else None

            # 遍历项目中的每一行数据
            for idx, (col_0, *cols_rest) in enumerate(project):
                item_id = col_0.split(":")[0]
                if primary_type in col_0:
                    primary_value = float(cols_rest[1])

                    valid_range = False
                    if operator == '=':
                        valid_range = primary_value == limit
                    elif operator == '>=':
                        # 确定上限，如果下一个限制存在，则使用下一个限制作为上限，否则为无穷大
                        upper_limit = next_limit_value if next_limit_value is not None else float('inf')
                        valid_range = limit <= primary_value < upper_limit
                    elif operator == '<=':
                        # 确定下限，如果前一个限制存在，则使用前一个限制作为下限，否则为无穷小
                        lower_limit = prev_limit_value if prev_limit_value is not None else float('-inf')
                        valid_range = lower_limit < primary_value <= limit

                    if valid_range:
                        filtered_ids.append(item_id)




            # 存储每个符合条件的USL和LSL值
            for col_0, *cols_rest in project:
                if col_0.split(":")[0] in filtered_ids and linked_type in col_0:
                    try:
                        # 尝试转换为浮点数
                        value = float(cols_rest[0]) / 2 if "位置度" in linked_type else float(cols_rest[0])
                        usl_value = float(cols_rest[1]) / 2 if "位置度" in linked_type else float(cols_rest[1])
                        # 假设LSL值为固定值
                        lsl_value = 0.000  
                        nominal_value = float(cols_rest[1]) / 2 if "位置度" in linked_type else float(cols_rest[1])
                        name = col_0
                    except ValueError:
                        # 如果转换失败，打印错误信息并跳过当前迭代
                        continue


                    # 更新所有项目中的最大值对应的USL、LSL和Nominal
                    if limit_index not in all_max_values or value > all_max_values[limit_index]:
                        all_max_usl_values[limit_index] = usl_value
                        all_max_lsl_values[limit_index] = lsl_value
                        all_max_nominal_values[limit_index] = nominal_value

                    # 更新每个项目中的最大值和对应名称
                    if limit_index not in max_values_per_project or value > max_values_per_project[limit_index]:
                        max_values_per_project[limit_index] = value
                        max_names_per_project[limit_index] = name

                    # 更新所有项目中的最大值和对应名称
                    if limit_index not in all_max_values or value > all_max_values[limit_index]:
                        all_max_values[limit_index] = value
                        all_max_names[limit_index] = name

        # 将每个项目的最大值写入Excel单元格
        for limit_index in range(len(range_limits)):
            max_value = max_values_per_project.get(limit_index, "")
            column_to_write = col_start - (6 - limit_index)
            max_value_cell = ws.cell(row=row_instrument + project_idx, column=column_to_write, value=max_value)

            # 设置单元格样式
            max_value_cell.alignment = center_alignment
            max_value_cell.font = Font(name='Arial', size=9, bold=True)
            max_value_cell.number_format = '0.000'
            max_value_cell.border = thin_border


        for idx, (col_0, *cols_rest) in enumerate(filtered_columns):
            col_index = col_start + idx
            if col_index > 3:
                row_to_check_empty = row_dimension + 7
                if ws.cell(row=row_to_check_empty, column=col_index).value is None:
                    for r in range(2, row_to_check_empty + 1):
                        src_cell = ws.cell(row=row_dimension + r, column=col_index - 1)
                        dest_cell = ws.cell(row=row_dimension + r, column=col_index)

                        # 复制公式和样式
                        if src_cell.value is not None:
                            if isinstance(src_cell.value, str) and '=' in src_cell.value:
                                old_col = get_column_letter(col_index - 1)
                                new_col = get_column_letter(col_index)
                                dest_cell.value = update_formula_column(src_cell.value, old_col, new_col)
                            else:
                                dest_cell.value = src_cell.value

                            # 复制样式
                            dest_cell.font = copy(src_cell.font)
                            dest_cell.border = copy(src_cell.border)
                            dest_cell.fill = copy(src_cell.fill)
                            dest_cell.number_format = src_cell.number_format
                            dest_cell.alignment = copy(src_cell.alignment)

            # 替换“位置度”为“偏移量”
            modified_col_0 = col_0.replace("位置度", "偏移量")

            cell_0 = ws.cell(row=row_dimension, column=col_start + idx, value=modified_col_0)
            cell_0.border = thin_border
            cell_0.alignment = center_alignment
            cell_0.font = font_dimension

            cell_1 = ws.cell(row=row_instrument + project_idx, column=col_index, value=float(cols_rest[0]))

            # 如果列标题是“位置度”，则对值除以2
            if "位置度" in col_0:
                modified_value = float(cols_rest[0]) / 2
                cell_1.value = modified_value  # 更新单元格的值
            else:
                cell_1.value = float(cols_rest[0])  # 使用原始值

            # 定义红色粗线边框样式
            red_thick_border = Border(left=Side(style='thick', color='FF0000'), 
                                      right=Side(style='thick', color='FF0000'), 
                                      top=Side(style='thick', color='FF0000'), 
                                      bottom=Side(style='thick', color='FF0000'))

            # 检查该单元格的值是否为项目最大值
            is_max_value = any(max_values_per_project.get(limit_index, None) == float(cols_rest[0]) / 2 for limit_index in range(len(range_limits)))
    
            if is_max_value:
                cell_1.font = Font(name='Arial', size=9, bold=True, color=Color(rgb='00FF0000'))  # 红色字体
                cell_1.border = red_thick_border  # 应用红色粗线边框
            else:
                cell_1.font = Font(name='Arial', size=9)  # 默认字体样式
                cell_1.border = thin_border  # 应用默认边框

            cell_1.alignment = Alignment(horizontal='center', vertical='center')  # 设置单元格内容居中
            cell_1.number_format = '0.000'


            # 处理 cell_2，如果列标题是“平面度”，则值固定为0
            if "平面度" in col_0:
                cell_2_value = 0.000  #值固定为0
            elif "平行度" in col_0:
                cell_2_value = 0.000  #值固定为0
            elif "垂直度" in col_0:
                cell_2_value = 0.000  #值固定为0
            elif "同心度" in col_0:
                cell_2_value = 0.000  #值固定为0
            elif "真圆度" in col_0:
                cell_2_value = 0.000  #值固定为0
            elif "位置度" in col_0:
                cell_2_value = float(cols_rest[1]) / 2
            else:
                cell_2_value = float(cols_rest[1])  # 否则使用cols_rest中的值

            cell_2 = ws.cell(row=row_nominal, column=col_start + idx, value=cell_2_value)
            cell_2.alignment = center_alignment
            cell_2.font = font_nominal
            cell_2.number_format = '0.000'
            cell_2.border = thin_border
            cell_2.fill = fill_2

            if "直径" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "间距" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "距离" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "半径" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "角度" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "坐标" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "短轴" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "长轴" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "数值" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            elif "位置度" in col_0:
                usl_value = float(cols_rest[1]) / 2
                lsl_value = 0.000
            elif "平面度" in col_0:
                usl_value = float(cols_rest[1])
                lsl_value = 0.000
            elif "平行度" in col_0:
                usl_value = float(cols_rest[1])
                lsl_value = 0.000
            elif "垂直度" in col_0:
                usl_value = float(cols_rest[1])
                lsl_value = 0.000
            elif "同心度" in col_0:
                usl_value = float(cols_rest[1])
                lsl_value = 0.000
            elif "真圆度" in col_0:
                usl_value = float(cols_rest[1])
                lsl_value = 0.000
            elif "最大间距" in col_0:
                usl_value = float(cols_rest[1]) + float(cols_rest[2])
                lsl_value = float(cols_rest[1]) - float(cols_rest[3])
            else:
                continue

            cell_usl = ws.cell(row=row_usl, column=col_start + idx, value=usl_value)
            cell_usl.alignment = center_alignment
            cell_usl.number_format = '0.000'
            cell_usl.border = thin_border

            cell_lsl = ws.cell(row=row_lsl, column=col_start + idx, value=lsl_value)
            cell_lsl.alignment = center_alignment
            cell_lsl.number_format = '0.000'
            cell_lsl.border = thin_border

    # 将所有项目的最大值对应的USL、LSL和Nominal值写入Excel单元格
    for limit_index, max_name in all_max_names.items():
        column_to_write = col_start - (6 - limit_index)

        # 获取最大值对应的USL、LSL和Nominal值
        max_usl_value = all_max_usl_values.get(limit_index, "无")
        max_lsl_value = all_max_lsl_values.get(limit_index, "无")
        max_nominal_value = all_max_nominal_values.get(limit_index, "无")

        # 写入名称、USL、LSL和Nominal值
        modified_max_name = max_name.replace("位置度", "偏移量").replace('"', '')
        max_name_cell = ws.cell(row=row_dimension, column=column_to_write, value=modified_max_name)
        max_usl_cell = ws.cell(row=row_usl, column=column_to_write, value=max_usl_value)
        max_lsl_cell = ws.cell(row=row_lsl, column=column_to_write, value=max_lsl_value)
        max_nominal_cell = ws.cell(row=row_nominal, column=column_to_write, value=max_nominal_value)

        # 设置单元格样式
        max_name_cell.alignment = center_alignment
        max_name_cell.font = Font(name='Arial', size=9, bold=True)
        max_name_cell.number_format = '0.000'
        max_name_cell.border = thin_border

        # 为USL、LSL和Nominal单元格设置样式，包括背景颜色
        for cell in [max_usl_cell, max_lsl_cell, max_nominal_cell]:
            cell.alignment = center_alignment
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.number_format = '0.000'
            cell.border = thin_border
            cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # 应用背景颜色样式




    wb = ws.parent  # 获取工作簿对象
    ws_second = wb.worksheets[second_sheet_index]  # 获取第二页工作表

    # 设置字体和边框样式
    font_name_type = Font(name='等线', size=11, bold=True, color='E5E2D1')
    font_standard_to_error = Font(name='Arial', size=11, color='E5E2D1')
    font_judgement = Font(name='Arial', size=11, bold=True, color='E5E2D1')
    thin_black_border = Border(left=Side(style='thin', color='000000'), 
                               right=Side(style='thin', color='000000'), 
                               top=Side(style='thin', color='000000'), 
                               bottom=Side(style='thin', color='000000'))

    # 设置对齐样式
    left_alignment = Alignment(horizontal='left')
    right_alignment = Alignment(horizontal='right')

    # 获取单元格位置信息
    type_cell_row, type_cell_col = cell_info.get('type', (None, None))
    name_cell_row, name_cell_col = cell_info.get('name', (None, None))
    standard_value_cell_col = cell_info.get('standard_value', (None, None))[1]
    measured_value_cell_col = cell_info.get('measured_value', (None, None))[1]
    error_cell_col = cell_info.get('error', (None, None))[1]
    upper_tolerance_cell_col = cell_info.get('upper_tolerance', (None, None))[1]  # 获取“正公差”列的位置
    lower_tolerance_cell_col = cell_info.get('lower_tolerance', (None, None))[1]  # 获取“负公差”列的位置
    judgement_cell_col = cell_info.get('judgement', (None, None))[1]  # 获取“判定”列的位置

    if type_cell_row is None or type_cell_col is None or name_cell_row is None or name_cell_col is None or standard_value_cell_col is None or measured_value_cell_col is None or error_cell_col is None or upper_tolerance_cell_col is None:
        print("Error: 未找到必要的单元格位置。")
        return

    # 首先收集所有“直径”和“偏移量”数据
    all_diameter_data = []
    all_position_data = []
    for project in projects:
        filtered_columns, prompt_data = process_project(project)
        input_name = prompt_data.get("输入", "")
    
        for col_0, *cols_rest in filtered_columns:
            col_0_modified = col_0.replace("位置度", "偏移量")
            if "直径" in col_0:
                all_diameter_data.append((input_name, col_0_modified, *cols_rest))
            elif "位置度" in col_0:
                all_position_data.append((input_name, col_0_modified, *cols_rest))

    # 然后按顺序写入所有“直径”数据，接着是所有“偏移量”数据
    current_row = type_cell_row + 1
    for data in all_diameter_data + all_position_data:
        input_name, col_0, *cols_rest = data

        col_0 = col_0.replace("位置度", "偏移量")

        # 根据列标题处理cell_2_value
        if "平面度" in col_0 or "平行度" in col_0 or "垂直度" in col_0 or "同心度" in col_0 or "真圆度" in col_0:
            cell_2_value = 0.000
        elif "偏移量" in col_0:
            cell_2_value = float(cols_rest[1]) / 2
        else:
            cell_2_value = float(cols_rest[1])

        if "直径" in col_0 or "偏移量" in col_0:
            # 写入“名称”和“类型”，设置对齐方式和边框
            name_cell = ws_second.cell(row=current_row, column=name_cell_col, value=input_name)
            type_cell = ws_second.cell(row=current_row, column=type_cell_col, value=col_0)
            name_cell.alignment = left_alignment
            type_cell.alignment = left_alignment
            name_cell.border = thin_black_border
            type_cell.border = thin_black_border
            name_cell.font = font_name_type
            type_cell.font = font_name_type

            # 写入“标准值”
            ws_second.cell(row=current_row, column=standard_value_cell_col, value=cell_2_value).number_format = '0.000'

            # 写入“测量值”，如果是“位置度”，则除以2
            measured_value = float(cols_rest[0]) / 2 if "偏移量" in col_0 else float(cols_rest[0])
            ws_second.cell(row=current_row, column=measured_value_cell_col, value=measured_value).number_format = '0.000'

            # 直接计算并写入“误差”
            error_value = measured_value - cell_2_value
            ws_second.cell(row=current_row, column=error_cell_col, value=error_value).number_format = '0.000'


            # 计算并写入“上公差”
            if "直径" in col_0:
                usl_value = float(cols_rest[2])
            elif "偏移量" in col_0:
                usl_value = float(cols_rest[1]) / 2 - cell_2_value
            ws_second.cell(row=current_row, column=upper_tolerance_cell_col, value=usl_value).number_format = '0.000'


            # 计算并写入“负公差”
            if "直径" in col_0:
                lsl_value = (float(cols_rest[1]) - float(cols_rest[3])) - cell_2_value
            elif "偏移量" in col_0:
                lsl_value = 0 - cell_2_value  # 对于“偏移量”，负公差为0减去标准值
            ws_second.cell(row=current_row, column=lower_tolerance_cell_col, value=lsl_value).number_format = '0.000'

            upper_tolerance = ws_second.cell(row=current_row, column=upper_tolerance_cell_col).value
            lower_tolerance = ws_second.cell(row=current_row, column=lower_tolerance_cell_col).value

            # 判定逻辑
            if error_value > upper_tolerance:  # 如果误差值超过正公差，直接标记为“Exceed”
                judgement = "Exceed"
            else:
                if upper_tolerance == 0 and lower_tolerance == 0:
                    judgement = "Good job" if error_value == 0 else "Check values"
                else:
                    error_percentage = abs(error_value) / max(abs(upper_tolerance), abs(lower_tolerance))
                    if error_percentage <= 0.10:
                        judgement = "Good job"
                    elif error_percentage > 1.00:
                        judgement = "Exceed"
                    else:
                        # 根据误差百分比接近100%的程度添加">"或"<"
                        error_close_to_100 = int((error_percentage - 0.10) / 0.18) + 1  # 至少1个，最多5个符号
                        if error_value > 0:
                            judgement = ">" * min(5, max(1, error_close_to_100))
                        else:
                            judgement = "<" * min(5, max(1, error_close_to_100))
            ws_second.cell(row=current_row, column=judgement_cell_col, value=judgement)

            # 设置除“名称”和“类型”以外的单元格的对齐方式和边框
            for col in range(standard_value_cell_col, judgement_cell_col + 1):
                cell = ws_second.cell(row=current_row, column=col)
                cell.alignment = right_alignment
                cell.border = thin_black_border
                if col in [standard_value_cell_col, measured_value_cell_col, error_cell_col, upper_tolerance_cell_col, lower_tolerance_cell_col]:
                    cell.font = font_standard_to_error
                if col == judgement_cell_col:
                    cell.font = font_judgement

            # 设置背景色的逻辑
            if error_value > usl_value:  # 测量值超过正公差
                # 设置超过正公差时的背景色
                fill_color_exceed_upper = '963634'  # 用于大部分单元格
                fill_color_exceed_critical = '752B29'  # 对于超正公差的标准值和测量值列使用更显眼的颜色

                # 应用通用背景色
                for col in range(name_cell_col, judgement_cell_col + 1):
                    if col in [standard_value_cell_col, measured_value_cell_col]:
                        # 对标准值和测量值列应用更显眼的背景色
                        ws_second.cell(row=current_row, column=col).fill = PatternFill(start_color=fill_color_exceed_critical, end_color=fill_color_exceed_critical, fill_type='solid')
                    else:
                        # 对其它单元格应用一般的超公差背景色
                        ws_second.cell(row=current_row, column=col).fill = PatternFill(start_color=fill_color_exceed_upper, end_color=fill_color_exceed_upper, fill_type='solid')


            elif error_percentage <= 1.00:
                # 如果误差百分比小于或等于100%，使用正常的颜色设置
                fill_color = '303030'
                ws_second.cell(row=current_row, column=standard_value_cell_col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                ws_second.cell(row=current_row, column=measured_value_cell_col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                fill_color_other = '404040'
                for col in range(name_cell_col, judgement_cell_col + 1):
                    if col not in [standard_value_cell_col, measured_value_cell_col]:
                        ws_second.cell(row=current_row, column=col).fill = PatternFill(start_color=fill_color_other, end_color=fill_color_other, fill_type='solid')
            else:
                # 误差超过100%，根据误差的正负使用不同的颜色
                fill_color = '752B29' if error_value > 0 else '770737'
                fill_color_other = '963634' if error_value > 0 else '9A0847'
                for col in range(name_cell_col, judgement_cell_col + 1):
                    if col in [standard_value_cell_col, measured_value_cell_col]:
                        ws_second.cell(row=current_row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    else:
                        ws_second.cell(row=current_row, column=col).fill = PatternFill(start_color=fill_color_other, end_color=fill_color_other, fill_type='solid')


            current_row += 1

    # 第三页工作表处理
    ws_third = wb.worksheets[third_sheet_index]

    # 获取第三页单元格位置信息
    name_cell_row_3rd, name_cell_col_3rd = cell_info_3rd.get('name', (None, None))
    type_cell_row_3rd, type_cell_col_3rd = cell_info_3rd.get('type', (None, None))
    x_standard_value_cell_col_3rd = cell_info_3rd.get('x_standard_value', (None, None))[1]
    y_standard_value_cell_col_3rd = cell_info_3rd.get('y_standard_value', (None, None))[1]
    x_measured_value_cell_col_3rd = cell_info_3rd.get('x_measured_value', (None, None))[1]
    y_measured_value_cell_col_3rd = cell_info_3rd.get('y_measured_value', (None, None))[1]
    x_error_cell_col_3rd = cell_info_3rd.get('x_error', (None, None))[1]
    y_error_cell_col_3rd = cell_info_3rd.get('y_error', (None, None))[1]
    diameter_cell_col_3rd = cell_info_3rd.get('diameter', (None, None))[1]
    concentricity_cell_col_3rd = cell_info_3rd.get('concentricity', (None, None))[1]
    concentricity_tolerance_cell_col_3rd = cell_info_3rd.get('concentricity_tolerance', (None, None))[1]

    # 设置字体和边框样式
    font_style = Font(name='等线', size=11, bold=True, color='E5E2D1')
    thin_black_border = Border(left=Side(style='thin', color='000000'), 
                               right=Side(style='thin', color='000000'), 
                               top=Side(style='thin', color='000000'), 
                               bottom=Side(style='thin', color='000000'))

    # 背景颜色设置
    fill_color_1 = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    fill_color_2 = PatternFill(start_color='343434', end_color='343434', fill_type='solid')

    # 筛选并合并包含“中心坐标”的数据
    combined_center_coordinate_data = {}
    diameter_data = {}  # 存储“直径”数据
    offset_data = {}  # 存储“位置度”即“偏移量”数据
    for project in projects:
        filtered_columns, prompt_data = process_project(project)
        input_name = prompt_data.get("输入", "")
        for col_0, *cols_rest in filtered_columns:
            if "中心坐标" in col_0:
                key = input_name + " " + col_0.split(":")[0]  # 如 "01 圆 8"
                if key not in combined_center_coordinate_data:
                    combined_center_coordinate_data[key] = {'input_name': input_name, 'type': col_0.split(":")[0], 'X_data': [], 'Y_data': []}
                if "X" in col_0 and len(cols_rest) > 1:
                    combined_center_coordinate_data[key]['X_data'] = cols_rest
                elif "Y" in col_0 and len(cols_rest) > 1:
                    combined_center_coordinate_data[key]['Y_data'] = cols_rest
            elif "直径" in col_0:
                diameter_key = input_name + " " + col_0.split(":")[0]  # 如 "01 圆 8"
                diameter_data[diameter_key] = cols_rest[1]  # 存储直径的值
            elif "位置度" in col_0:
                offset_key = input_name + " " + col_0.split(":")[0]
                offset_data[offset_key] = float(cols_rest[1]) / 2  # 存储偏移量的值

    # 创建一个列表来存储被删除的行的数据
    deleted_rows_data = []

    # 对数据进行分组和筛选
    grouped_data = {}
    for key, value in combined_center_coordinate_data.items():
        input_name = value['input_name']
        # 确保X和Y数据都存在并且长度一致
        if len(value['X_data']) > 1 and len(value['Y_data']) > 1:
            x_standard_value = float(value['X_data'][1])
            y_standard_value = float(value['Y_data'][1])
            group_key = (input_name, x_standard_value, y_standard_value)  # 使用X和Y的标准值进行分组
            if group_key not in grouped_data:
                grouped_data[group_key] = []
            grouped_data[group_key].append(value)

    # 只保留至少有一对匹配的X和Y数据的组
    filtered_grouped_data = {}
    for k, v in grouped_data.items():
        if len(v) > 1:
            filtered_grouped_data[k] = v
        else:
            deleted_rows_data.extend(v)  # 将没有匹配的行保存在deleted_rows_data列表中

    # 按照X和Y标准值大小进行排序
    sorted_group_keys = sorted(filtered_grouped_data.keys(), key=lambda x: (x[0], x[1], x[2]))

    # 写入合并后的“中心坐标”数据
    current_row_3rd = name_cell_row_3rd + 1
    use_first_color = True  # 初始时使用第一种颜色
    for group_key in sorted_group_keys:
        group_values = filtered_grouped_data[group_key]
        max_concentricity_tolerance = max([offset_data.get(value['input_name'] + " " + value['type'], 0) for value in group_values])  # 获取最大同心度公差
        exceed_tolerance = False  # 初始化同心度是否超过公差的标志

        # 计算同心度，检查是否超过公差
        if len(group_values) > 1:
            value = group_values[0]
            next_value = group_values[1]
            x_distance = float(value['X_data'][0]) - float(next_value['X_data'][0])
            y_distance = float(value['Y_data'][0]) - float(next_value['Y_data'][0])
            concentricity = math.sqrt(x_distance**2 + y_distance**2)
            if concentricity > max_concentricity_tolerance:
                exceed_tolerance = True  # 同心度超过公差

        for i in range(len(group_values)):
            value = group_values[i]
            # 写入“名称”和“类型”
            ws_third.cell(row=current_row_3rd, column=name_cell_col_3rd, value=value['input_name'])
            ws_third.cell(row=current_row_3rd, column=type_cell_col_3rd, value=value['type'] + ": X中心坐标, Y中心坐标")

            # 写入X中心坐标的标准值
            cell_x_standard = ws_third.cell(row=current_row_3rd, column=x_standard_value_cell_col_3rd, value=group_key[1])
            cell_x_standard.number_format = '0.000'

            # 写入X中心坐标的测量值
            cell_x_measured = ws_third.cell(row=current_row_3rd, column=x_measured_value_cell_col_3rd, value=float(value['X_data'][0]))
            cell_x_measured.number_format = '0.000'

            # 写入Y中心坐标的标准值
            cell_y_standard = ws_third.cell(row=current_row_3rd, column=y_standard_value_cell_col_3rd, value=group_key[2])
            cell_y_standard.number_format = '0.000'

            # 写入Y中心坐标的测量值
            cell_y_measured = ws_third.cell(row=current_row_3rd, column=y_measured_value_cell_col_3rd, value=float(value['Y_data'][0]))
            cell_y_measured.number_format = '0.000'

            # 计算并写入X误差
            x_error = float(value['X_data'][0]) - group_key[1]
            cell_x_error = ws_third.cell(row=current_row_3rd, column=x_error_cell_col_3rd, value=x_error)
            cell_x_error.number_format = '0.000'

            # 计算并写入Y误差
            y_error = float(value['Y_data'][0]) - group_key[2]
            cell_y_error = ws_third.cell(row=current_row_3rd, column=y_error_cell_col_3rd, value=y_error)
            cell_y_error.number_format = '0.000'

            # 查找并写入对应的“直径”值
            diameter_key = value['input_name'] + " " + value['type']
            if diameter_key in diameter_data:
                diameter_value = float(diameter_data[diameter_key])
                cell_diameter = ws_third.cell(row=current_row_3rd, column=diameter_cell_col_3rd, value=diameter_value)
            else:
                # 如果没有找到对应的“直径”值，可以选择写入默认值或者留空
                cell_diameter = ws_third.cell(row=current_row_3rd, column=diameter_cell_col_3rd, value=None)  # 或者设置默认值，例如0.000
            cell_diameter.number_format = '0.000'

            # 计算同心度并写入（只在第一个匹配项中添加同心度）
            if i == 0 and len(group_values) > 1:
                next_value = group_values[1]  # 取第二个匹配项
                x_distance = float(value['X_data'][0]) - float(next_value['X_data'][0])
                y_distance = float(value['Y_data'][0]) - float(next_value['Y_data'][0])
                concentricity = math.sqrt(x_distance**2 + y_distance**2)
                cell_concentricity = ws_third.cell(row=current_row_3rd, column=concentricity_cell_col_3rd, value=concentricity)
                cell_concentricity.number_format = '0.000'

                # 写入同心度公差（最大值）
                cell_concentricity_tolerance = ws_third.cell(row=current_row_3rd, column=concentricity_tolerance_cell_col_3rd, value=max_concentricity_tolerance)
                cell_concentricity_tolerance.number_format = '0.000'
            else:
                # 其他行同心度和同心度公差保留空白
                cell_concentricity = ws_third.cell(row=current_row_3rd, column=concentricity_cell_col_3rd, value=None)
                cell_concentricity_tolerance = ws_third.cell(row=current_row_3rd, column=concentricity_tolerance_cell_col_3rd, value=None)

            # 选择背景颜色
            current_fill_color = fill_color_1 if use_first_color else fill_color_2

            # 为每个单元格设置样式
            for col in range(name_cell_col_3rd, concentricity_cell_col_3rd + 1):
                cell = ws_third.cell(row=current_row_3rd, column=col)
                cell.font = font_style
                cell.border = thin_black_border

                # 根据同心度是否超过公差设置背景颜色
                if exceed_tolerance:
                    cell.fill = PatternFill(start_color='963634', end_color='963634', fill_type='solid')
                else:
                    cell.fill = current_fill_color

            current_row_3rd += 1

        # 在处理完每个匹配项后改变颜色
        use_first_color = not use_first_color

    # 在匹配行后添加没有匹配的单行数据
    for value in deleted_rows_data:
        # 写入“名称”和“类型”
        ws_third.cell(row=current_row_3rd, column=name_cell_col_3rd, value=value['input_name'])
        ws_third.cell(row=current_row_3rd, column=type_cell_col_3rd, value=value['type'] + ": X中心坐标, Y中心坐标")

        # 写入X中心坐标的标准值
        if len(value['X_data']) > 1:
            cell_x = ws_third.cell(row=current_row_3rd, column=x_standard_value_cell_col_3rd, value=float(value['X_data'][1]))
            cell_x.number_format = '0.000'

        # 写入Y中心坐标的标准值
        if len(value['Y_data']) > 1:
            cell_y = ws_third.cell(row=current_row_3rd, column=y_standard_value_cell_col_3rd, value=float(value['Y_data'][1]))
            cell_y.number_format = '0.000'

        # 写入X中心坐标的测量值
        if len(value['X_data']) > 1:
            cell_x_measured = ws_third.cell(row=current_row_3rd, column=x_measured_value_cell_col_3rd, value=float(value['X_data'][0]))
            cell_x_measured.number_format = '0.000'

        # 写入Y中心坐标的测量值
        if len(value['Y_data']) > 1:
            cell_y_measured = ws_third.cell(row=current_row_3rd, column=y_measured_value_cell_col_3rd, value=float(value['Y_data'][0]))
            cell_y_measured.number_format = '0.000'

        # 计算并写入X误差
        if len(value['X_data']) > 1:
            x_error = float(value['X_data'][0]) - float(value['X_data'][1])
            cell_x_error = ws_third.cell(row=current_row_3rd, column=x_error_cell_col_3rd, value=x_error)
            cell_x_error.number_format = '0.000'

        # 计算并写入Y误差
        if len(value['Y_data']) > 1:
            y_error = float(value['Y_data'][0]) - float(value['Y_data'][1])
            cell_y_error = ws_third.cell(row=current_row_3rd, column=y_error_cell_col_3rd, value=y_error)
            cell_y_error.number_format = '0.000'

        # 写入“直径”值
        diameter_key = value['input_name'] + " " + value['type']
        if diameter_key in diameter_data:
            diameter_value = float(diameter_data[diameter_key])
            cell_diameter = ws_third.cell(row=current_row_3rd, column=diameter_cell_col_3rd, value=diameter_value)
        else:
            # 如果没有找到对应的“直径”值，可以选择写入默认值或者留空
            cell_diameter = ws_third.cell(row=current_row_3rd, column=diameter_cell_col_3rd, value=None)  # 或者设置默认值，例如0.000
        cell_diameter.number_format = '0.000'

        # 为每个单元格设置样式（字体、边框、背景颜色）
        for col in range(name_cell_col_3rd, concentricity_cell_col_3rd + 1):  # 从名称列到同心度列
            cell = ws_third.cell(row=current_row_3rd, column=col)
            cell.font = font_style
            cell.border = thin_black_border
            cell.fill = fill_color_1

        current_row_3rd += 1


    # 获取第四页工作表对象
    ws_fourth = wb.worksheets[fourth_sheet_index]

    # 确定“类型”单元格的起始行和各列位置
    type_cell_row, type_cell_col = cell_info_4th.get('type', (None, None))
    diameter_col = cell_info_4th.get('diameter', (None, None))[1]
    standard_value_col = cell_info_4th.get('standard_value', (None, None))[1]
    measured_value_col = cell_info_4th.get('measured_value', (None, None))[1]

    if type_cell_row is None or type_cell_col is None or diameter_col is None or measured_value_col is None:
        print("Error: 未找到第四页必要的'类型'、'直径'或'测量值'单元格位置。")
        return

    # 初始化存储每个类型的测量值数量的字典
    type_measurement_counts = {}

    # 初始化一个字典来存储每个类型对应的行号
    type_row_mapping = {}

    # 初始化变量以跟踪当前应写入的行
    current_row = type_cell_row + 1

    # 假设我们有一个新的字典来直接映射列号到input_name
    column_to_input_name_map = {}

    # 遍历项目数据，写入“偏移量”类型信息
    for data in all_position_data:
        input_name, col_0, measure_value, standard_value, *_ = data
        type_name = col_0.split(":")[0].strip()

        if type_name not in type_row_mapping:
            type_row_mapping[type_name] = current_row
            current_row += 1  # 为新类型准备新行
            type_measurement_counts[type_name] = 0

        # 直接将当前处理的列与input_name关联

        row_for_type = type_row_mapping[type_name]

        # 计算当前类型的测量值应该写入的列
        current_measured_value_col = measured_value_col + type_measurement_counts[type_name]

        column_to_input_name_map[current_measured_value_col] = input_name

        if type_measurement_counts[type_name] == 0:
            # 写入“类型”，并应用样式
            type_cell = ws_fourth.cell(row=row_for_type, column=type_cell_col, value=col_0)
            type_cell.font = font_name_type
            type_cell.border = thin_black_border
            type_cell.alignment = left_alignment  # “类型”单元格左对齐

            # 查找相同名称的直径数据的标准值并写入
            diameter_standard_value = None
            for diameter_data in all_diameter_data:
                if diameter_data[0] == input_name and diameter_data[1].startswith(type_name):
                    diameter_standard_value = float(diameter_data[3])
                    break
            # 查找相同名称的直径数据的标准值并写入
            if diameter_standard_value is not None and diameter_col is not None:
                diameter_cell = ws_fourth.cell(row=row_for_type, column=diameter_col, value=diameter_standard_value)
                diameter_cell.font = font_name_type
                diameter_cell.border = thin_black_border
                diameter_cell.alignment = right_alignment  # 设置为右对齐
                diameter_cell.number_format = '0.000'

            # 查找偏移量的标准值并写入到“标准值”列
            offset_standard_value = None
            for position_data in all_position_data:
                if position_data[0] == input_name and position_data[1].startswith(type_name):
                    offset_standard_value = float(position_data[3]) / 2
                    break
            if offset_standard_value is not None and standard_value_col is not None:
                standard_value_cell = ws_fourth.cell(row=row_for_type, column=standard_value_col, value=offset_standard_value)
                standard_value_cell.font = font_name_type
                standard_value_cell.border = thin_black_border
                standard_value_cell.alignment = right_alignment
                standard_value_cell.number_format = '0.000'

        # 查找偏移量的测量值并写入到相应的列，同时除以2
        offset_measured_value = None
        for position_data in all_position_data:
            if position_data[0] == input_name and position_data[1].startswith(type_name):
                offset_measured_value = float(position_data[2]) / 2
                break
        if offset_measured_value is not None:
            # 写入测量值
            measured_value_cell = ws_fourth.cell(row=row_for_type, column=current_measured_value_col, value=float(measure_value) / 2)
            measured_value_cell.font = font_name_type
            measured_value_cell.border = thin_black_border
            measured_value_cell.alignment = right_alignment
            measured_value_cell.number_format = '0.0000'

        # 更新该类型的测量值计数
        type_measurement_counts[type_name] += 1

    # 创建散点图
    chart = ScatterChart()
    chart.title = filename_without_extension
    chart.style = 2  # 使用预定义的样式
    chart.x_axis.title = '针点位置'
    chart.y_axis.title = '偏移量'

    # 确定整个测量值区域的起始行和结束行
    start_row = type_cell_row + 1
    end_row = current_row - 1

    # 确定整个测量值区域的起始列和结束列
    start_col = measured_value_col
    end_col = start_col  # 初始假设为起始列

    # 遍历所有类型，更新测量值的结束列
    for type_name, counts in type_measurement_counts.items():
        if counts > 0:
            end_col = max(end_col, start_col + counts - 1)

    # 创建一个辅助列来存放递增的序列，不隐藏，样式与其他单元格一致
    aux_col = end_col + 1
    for i, row_num in enumerate(range(start_row, end_row + 1), start=0):
        aux_cell = ws_fourth.cell(row=row_num, column=aux_col, value=i)
        aux_cell.font = font_name_type
        aux_cell.border = thin_black_border
        aux_cell.alignment = right_alignment  # 辅助列也设置为右对齐

    # 然后在添加散点图系列时使用column_to_input_name_map来获取每个系列的标题
    for col in range(start_col, end_col + 1):
        input_id = column_to_input_name_map.get(col, "Unknown")

        x_values = Reference(ws_fourth, min_col=aux_col, min_row=start_row, max_col=aux_col, max_row=end_row)
        y_values = Reference(ws_fourth, min_col=col, min_row=start_row, max_col=col, max_row=end_row)

        series = Series(values=y_values, xvalues=x_values, title=f"{input_id}")

    
        # 设置系列的标记样式为无，以隐藏点
        series.marker = Marker(symbol='none')
    
        # 设置系列为平滑的线条，注意：这可能不会按预期工作，取决于OpenPyXL的支持
        series.smooth = True  # 尝试使线条平滑

        chart.series.append(series)

    chart_row = start_row + 2  # 空出两行的起始位置
    chart_col = aux_col + 2  # 留一列空白后添加图表
    chart_cell = ws_fourth.cell(row=chart_row, column=chart_col).coordinate

    # 尝试调整网格线设置（注意：openpyxl可能不支持仅显示横线的精细控制）
    chart.x_axis.majorGridlines = None  # 不显示X轴的主要网格线
    chart.y_axis.majorGridlines = ChartLines()  # 显示Y轴的主要网格线

    # 设置图例位置
    chart.legend.position = 't'  # 't' 代表顶部（top）

    ws_fourth.add_chart(chart, chart_cell)



    # 调整散点图的大小
    chart.width = 35  # 散点图的宽度，单位为英寸
    chart.height = 15  # 散点图的高度，单位为英寸

    # 定义背景色填充样式
    fill_measurements = PatternFill(start_color='303030', end_color='303030', fill_type='solid')  # 测量值列的背景色
    fill_others = PatternFill(start_color='404040', end_color='404040', fill_type='solid')  # 其他列的背景色
    fill_aux_after = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 辅助列之后的背景色

    # 应用背景色填充到所有相关单元格
    for row in ws_fourth.iter_rows(min_row=type_cell_row + 1, max_row=current_row - 1, min_col=type_cell_col, max_col=aux_col + 28):
        for cell in row:
            # 根据列号决定填充颜色
            if start_col <= cell.column <= end_col:
                cell.fill = fill_measurements
            elif cell.column > aux_col and cell.column <= aux_col + 28:
                cell.fill = fill_aux_after
            else:
                cell.fill = fill_others






class ConversionThread(QThread):
    conversion_done = pyqtSignal()
    file_converted = pyqtSignal(str)  # 发射信号，参数为转换完成的文件名
    is_running = False  # 用于检查线程是否已在运行
    initial_folder_info = {}  # 用于记录初始文件夹的信息

    def __init__(self, template_path, folder_path, file_hashes, judgement_type, selected_method, range_limit, machine_number, consider_xy_center):
        super().__init__()

        self.judgement_type = judgement_type
        self.selected_method = selected_method
        self.range_limit = range_limit
        self.machine_number = machine_number

        self.day_shift_start = None
        self.day_shift_end = None
        self.night_shift_start = None
        self.night_shift_end = None
        print("ConversionThread initialized with default shift times")

        self.consider_xy_center = consider_xy_center  # 新增属性，存储XY中心坐标的考虑状态
        self.template_path = template_path
        self.folder_path = folder_path  
        self.file_hashes = file_hashes
        self.last_col_index = 1  # 初始化为第一列

    def update_shift_times(self, day_start, day_end, night_start, night_end):
        self.day_shift_start = day_start
        self.day_shift_end = day_end
        self.night_shift_start = night_start
        self.night_shift_end = night_end
        print(f"Shift times updated: Day Shift {day_start.toString()} - {day_end.toString()}, Night Shift {night_start.toString()} - {night_end.toString()}")

    def get_image_count_for_folder(self, folder_name):
        """获取与文件夹名称相关的图片数量"""
        image_count = 0
        for file_name in os.listdir(self.folder_path):
            if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')) and folder_name in file_name:
                image_count += 1
        return image_count

    def run(self):
        if ConversionThread.is_running:
            print("线程已在运行。")
            return

        ConversionThread.is_running = True

        # 检查文件夹是否存在
        if not os.path.exists(self.folder_path):
            print(f"文件夹 {self.folder_path} 不存在或已被移除。")
            ConversionThread.is_running = False
            return

        current_folder_info = {}  # 记录当前文件夹的信息


        for file_name in os.listdir(self.folder_path):
            file_path = os.path.join(self.folder_path, file_name)
            # 每处理一个文件夹前，重置列索引
            self.last_col_index = 1

            if os.path.isfile(file_path) and file_name.lower().endswith('.txt'):
                try:
                    txt_path = os.path.join(self.folder_path, file_name)
                    with open(txt_path, 'rb') as f:
                        file_hash = md5(f.read()).hexdigest()

                    if file_name.lower().endswith('.txt') and (file_name not in self.file_hashes or self.file_hashes[file_name] != file_hash):
                        wb = load_workbook(self.template_path)
                        ws = wb.active
                        rows = find_rows(ws)
                        row_diagram = rows[-2]  # 获取“测量示意图”所在行
                        row_dimension = rows[0]  # 获取“Dimension”所在行
                        lines = read_txt_lines(txt_path)
                        projects = process_lines(lines)

                        # 获取最早时间和班次
                        earliest_time, shift = self.get_latest_time_and_shift(projects)

                        # 获取文件名不带扩展名
                        filename_without_extension = os.path.splitext(file_name)[0]

                        wb = load_workbook(self.template_path)
                        ws = wb.active
                        second_sheet_index = wb.index(wb.worksheets[1])  # 获取第二页工作表索引
                        third_sheet_index = wb.index(wb.worksheets[2])  # 获取第三页工作表索引
                        fourth_sheet_index = wb.index(wb.worksheets[3])  # 获取第四页工作表索引

                        # 获取第二页和第三页所需单元格的位置
                        second_sheet = wb.worksheets[second_sheet_index]
                        third_sheet = wb.worksheets[third_sheet_index]
                        fourth_sheet = wb.worksheets[fourth_sheet_index]

                        cell_info = find_name_cell(second_sheet)
                        cell_info_3rd = find_third_sheet_cells(third_sheet)
                        cell_info_4th = find_name_cell_page4(fourth_sheet)

                        # 调用write_to_excel并传递所有信息
                        write_to_excel(ws, projects, rows, filename_without_extension, self.judgement_type, self.selected_method, self.range_limit, self.machine_number, self.consider_xy_center, second_sheet_index, cell_info, third_sheet_index, cell_info_3rd, fourth_sheet_index, cell_info_4th)

                        # 创建与文件同名的文件夹
                        new_folder_path = os.path.join(self.folder_path, filename_without_extension)
                        if not os.path.exists(new_folder_path):
                            os.makedirs(new_folder_path)

                        # 清理文件名，并添加最早时间和班次
                        earliest_time_str = earliest_time.strftime('%Y-%m-%d') if earliest_time else ''
                        cleaned_name = clean_filename(f"{self.machine_number}# {filename_without_extension} {earliest_time_str} {shift}")
                        output_name = cleaned_name + '.xlsx'
                        output_path = os.path.join(self.folder_path, output_name)  # 保存到原始文件夹

                        wb.save(output_path)
                        self.file_hashes[file_name] = file_hash  # 更新哈希
                        self.file_converted.emit(file_name)  # 发射转换完成的信号
                        self.conversion_done.emit()  # 发出转换完成信号

                        # 检查是否有图片
                        self.check_and_insert_image(new_folder_path, output_path, row_diagram, row_dimension)

                        # 转换完成后，记录初始文件夹信息
                        new_folder_name = os.path.splitext(file_name)[0]
                        new_folder_path = os.path.join(self.folder_path, new_folder_name)
                        if os.path.exists(new_folder_path):
                            print(f"记录文件夹 {new_folder_name} 的初始状态。")
                            initial_info = self.record_folder_info(new_folder_path)
                            self.initial_folder_info[new_folder_path] = initial_info
                            print(f"文件夹 {new_folder_name} 的初始状态: {initial_info}")

                except Exception as e:
                    print(f"处理文件 {file_name} 时发生错误: {e}")

        # 实时更新并检查当前文件夹信息
        for folder_path in self.initial_folder_info.keys():
            current_info = self.record_folder_info(folder_path)
            current_folder_info[folder_path] = current_info
            print(f"文件夹 {os.path.basename(folder_path)} 的实时状态: {current_info}")
            if current_info != self.initial_folder_info[folder_path]:
                print(f"检测到文件夹 {os.path.basename(folder_path)} 的变动，更新文件转换记录。")
                self.remove_converted_files(folder_path)

        # 检查并移动相似图片
        self.move_similar_images()

        ConversionThread.is_running = False

    def move_similar_images(self):
        """将包含关键字的图片移动到对应的文件夹中"""
        for folder_name in os.listdir(self.folder_path):
            folder_path = os.path.join(self.folder_path, folder_name)
            if os.path.isdir(folder_path):
                for file_name in os.listdir(self.folder_path):
                    if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                        # 检查文件名是否包含文件夹名称的关键字
                        keyword = file_name.split('.')[0]
                        if keyword in folder_name:
                            src_path = os.path.join(self.folder_path, file_name)
                            dest_path = os.path.join(folder_path, file_name)

                            # 如果目标文件夹中已存在同名文件，则添加序号
                            counter = 1
                            base_name, ext = os.path.splitext(file_name)
                            while os.path.exists(dest_path):
                                new_name = f"{base_name}_{counter}{ext}"
                                dest_path = os.path.join(folder_path, new_name)
                                counter += 1

                            shutil.move(src_path, dest_path)
                            print(f"图片 {file_name} 已移动到文件夹 {folder_name}")

    def record_folder_info(self, folder_path):
        """记录文件夹内的文件数量和哈希值"""
        file_count = 0
        file_hashes = {}

        if not os.path.exists(folder_path):
            print(f"文件夹不存在: {folder_path}")
            self.remove_converted_files(folder_path)  # 删除相关文件记录
            return {'file_count': file_count, 'file_hashes': file_hashes}

        for file_name in os.listdir(folder_path):
            if file_name.startswith('~$'):
                continue  # 跳过以 '~$' 开头的临时文件

            file_path = os.path.join(folder_path, file_name)
            if os.path.isfile(file_path):
                file_count += 1
                file_hash = self.get_file_hash(file_path)
                file_hashes[file_name] = file_hash

        return {'file_count': file_count, 'file_hashes': file_hashes}

    def get_file_hash(self, file_path):
        """计算文件的哈希值"""
        hash_md5 = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def remove_converted_files(self, folder_path):
        """从已转换文件名列表中删除对应文件夹内的所有文件记录"""
        folder_name = os.path.basename(folder_path)
        file_names_to_remove = [fn for fn in self.file_hashes if fn.startswith(folder_name)]
        for file_name in file_names_to_remove:
            del self.file_hashes[file_name]
        print(f"已删除文件夹 {folder_name} 的转换记录")

    def check_and_insert_image(self, folder_path, excel_file_path, row_diagram, row_dimension):
        print(f"检查文件夹 {folder_path} 中的图片")
        found_image = False
        for img_file in os.listdir(folder_path):
            if img_file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                found_image = True
                img_path = os.path.join(folder_path, img_file)
                print(f"找到图片 {img_path}，正在插入到 Excel")
                self.insert_image_to_excel(excel_file_path, img_path, row_diagram, row_dimension)
                # 如果您想处理所有找到的图片，删除 break 语句
                # break
        if not found_image:
            print(f"在文件夹 {folder_path} 中未找到图片")


    def is_time_in_range(self, start, end, check_time):
        if start <= end:
            return start <= check_time <= end
        else:  # 跨越午夜的情况
            return check_time >= start or check_time <= end

    def get_latest_time_and_shift(self, projects):
        day_shift_start = self.day_shift_start.toPyTime()
        day_shift_end = self.day_shift_end.toPyTime()
        night_shift_start = self.night_shift_start.toPyTime()
        night_shift_end = self.night_shift_end.toPyTime()

        latest_time = None
        shift = '白班'  # 默认设置为白班

        for project in projects:
            for line in project:
                if '日期/时间' in line[0]:
                    date_time_str = line[1].replace('"', '').strip()
                    try:
                        date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
                        if latest_time is None or date_time_obj > latest_time:
                            latest_time = date_time_obj
                            time_obj = latest_time.time()
                            if self.is_time_in_range(day_shift_start, day_shift_end, time_obj):
                                shift = '白班'
                            elif self.is_time_in_range(night_shift_start, night_shift_end, time_obj):
                                shift = '夜班'
                    except ValueError as e:
                        print(f"解析日期时间时发生错误: {e}，原始字符串: '{date_time_str}'")
                        continue  # 继续尝试解析下一行

        return latest_time, shift

    def insert_image_to_excel(self, excel_file_path, img_path, row_diagram, row_dimension):
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # 使用 PIL 加载图片来获取其尺寸
        pil_img = PILImage.open(img_path)
        orig_width, orig_height = pil_img.size

        # 计算插入区域的总高度
        total_height_points = sum(ws.row_dimensions[row].height for row in range(row_diagram + 1, row_dimension - 1) if ws.row_dimensions[row].height)
        total_height_pixels = total_height_points * 1.25  # 行高点数转换为像素

        # 计算缩放后的图片大小
        ratio = total_height_pixels / orig_height
        new_width = round(orig_width * ratio)
        new_height = round(orig_height * ratio)

        # 创建图片对象
        img = OpenpyxlImage(img_path)
        img.width, img.height = new_width, new_height

        # 确定图片的插入列
        start_col_index = self.last_col_index

        # 计算图片占用的列数
        col_span = self.calculate_columns_spanned(ws, new_width, start_col_index)

        # 设置图片插入位置
        img.anchor = get_column_letter(start_col_index) + str(row_diagram + 1)
        ws.add_image(img)

        # 更新插入后的最后列索引
        self.last_col_index = start_col_index + col_span

        # 保存工作簿
        wb.save(excel_file_path)
        print(f"图片已插入到 {excel_file_path} 在 {img.anchor}")

    def calculate_columns_spanned(self, ws, img_width_pixels, start_col_index):
        """计算图片占用的列数"""
        total_width_pixels = 0
        col_count = 0
        for col in range(start_col_index, ws.max_column + 1):
            col_width = ws.column_dimensions[get_column_letter(col)].width
            if col_width is None:
                continue  # 跳过未定义宽度的列
            col_width_pixels = col_width * 7  # 列宽转换为像素
            total_width_pixels += col_width_pixels
            col_count += 1
            if total_width_pixels >= img_width_pixels:
                break
        return col_count

    def calculate_next_col_index(self, ws, previous_col_index, col_span):
        """根据前一张图片的位置和占用列数计算下一张图片的起始列索引"""
        return previous_col_index + col_span   
            

class Toast(QWidget):
    def __init__(self, message, width=200, height=60, parent=None):
        super(Toast, self).__init__(parent)
        self.setFixedSize(width, height)  # 设置固定大小
        self.setAttribute(Qt.WA_DeleteOnClose)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setStyleSheet("background-color: white; color: black; border-radius: 7px;")

        # 创建布局并将 label 添加到布局中
        layout = QVBoxLayout(self)
        self.label = QLabel(message, self)
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)
        self.setLayout(layout)

    def text(self):
        # 添加一个新方法来获取 label 的文本
        return self.label.text()


    def show_toast(self, duration=4000):
        self.show()
        QTimer.singleShot(duration, self.close)


class CustomValidator(QValidator):
    def validate(self, string, pos):
        if all(c in "0123456789.<>=" for c in string):
            return QValidator.Acceptable, string, pos
        return QValidator.Invalid, string, pos


class ColorBlockWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.animation_position = 0
        self.time = 0  # 添加这行来初始化time属性
        self.alpha = 1.0  # 初始化alpha值为1
        self.speed_factor = uniform(0.005, 0.02)  # 随机速度因子，每次变化的速度在0.005到0.02之间
        self.gradient_colors = [QColor('#36D1DC'), QColor('#5B86E5')]  # 默认呼吸渐变色
        self.is_breathing = True  # 控制动画模式：True为呼吸渐变，False为滚动渐变

        # 呼吸渐变的定时器
        self.breathing_timer = QTimer(self)
        self.breathing_timer.timeout.connect(self.animateBreathing)
        self.breathing_timer.start(20)  # 呼吸渐变更新频率

        # 滚动渐变的定时器
        self.scrolling_timer = QTimer(self)
        self.scrolling_timer.timeout.connect(self.animateScrolling)
        # 滚动渐变的定时器默认不启动，只有在滚动渐变模式下才启动
        self.scrolling_timer_interval = 7  # 滚动渐变更新频率

    def setGradientColors(self, start_color, end_color, is_breathing=True):
        self.gradient_colors = [QColor(start_color), QColor(end_color)]
        self.is_breathing = is_breathing
        if not is_breathing:
            self.gradient_colors.append(QColor(start_color))  # 为滚动效果添加第三个颜色
            self.breathing_timer.stop()
            self.scrolling_timer.start(self.scrolling_timer_interval)
        else:
            self.scrolling_timer.stop()
            self.breathing_timer.start()
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # 设置画笔为无边框
        painter.setPen(Qt.NoPen)

        if self.is_breathing:
            # 纯色循环平滑切换逻辑
            current_color_index = int(self.animation_position) % len(self.gradient_colors)
            next_color_index = (current_color_index + 1) % len(self.gradient_colors)
            current_color = self.gradient_colors[current_color_index].toRgb()
            next_color = self.gradient_colors[next_color_index].toRgb()

            # 使用alpha混合计算当前颜色
            mixed_color = QColor.fromRgbF(
                (1 - self.alpha) * current_color.redF() + self.alpha * next_color.redF(),
                (1 - self.alpha) * current_color.greenF() + self.alpha * next_color.greenF(),
                (1 - self.alpha) * current_color.blueF() + self.alpha * next_color.blueF()
            )
            painter.setBrush(mixed_color)
            painter.drawRoundedRect(self.rect(), 3, 3)  # 绘制圆角矩形，角半径为2
        else:
            # 滚动渐变逻辑
            gradient = QLinearGradient(self.animation_position - self.width(), 0, self.animation_position, 0)
            for i, color in enumerate(self.gradient_colors):
                gradient.setColorAt(i / (len(self.gradient_colors) - 1), color)
            painter.setBrush(gradient)
            painter.drawRoundedRect(self.rect(), 3, 3)  # 绘制圆角矩形，角半径为2


    def animateBreathing(self):
        self.alpha -= self.speed_factor  # 根据速度因子更新alpha值
        if self.alpha <= 0:
            self.animation_position = (self.animation_position + 1) % len(self.gradient_colors)  # 循环动画位置
            self.alpha = 1.0  # 重置alpha值
            self.speed_factor = uniform(0.005, 0.02)  # 为下一个周期随机选择新的速度因子
        self.update()

    def animateScrolling(self):
        self.animation_position += 2
        if self.animation_position >= 2 * self.width():
            self.animation_position = 0
        self.update()


class UpdateCheckerThread(QThread):
    # 新版本号，程序下载链接，版本描述，数据模板版本号，数据模板下载链接
    update_available = pyqtSignal(str, str, str, str, str)
    update_check_finished = pyqtSignal()
    update_check_failed = pyqtSignal(str)

    def __init__(self, api_url, current_version):
        super().__init__()
        self.api_url = api_url
        self.current_version = current_version

    def run(self):
        try:
            response = requests.get(self.api_url)
            response.raise_for_status()
            latest_release = response.json()
            latest_version = latest_release['tag_name']
            description = latest_release['body']
            program_download_url = None
            data_file_version = None
            data_file_download_url = None

            exe_pattern = re.compile(r'.exe$', re.IGNORECASE)
            data_pattern = re.compile(r'数据模板(\d+\.\d+)|v(\d+\.\d+)\.xlsx', re.IGNORECASE)

            for asset in latest_release.get('assets', []):
                if exe_pattern.search(asset['name']):
                    program_download_url = asset['browser_download_url']
                else:
                    match = data_pattern.search(asset['name'])
                    if match:
                        version = match.group(1) or match.group(2)
                        if not data_file_version or self.is_newer_version(version, data_file_version):
                            data_file_version = version
                            data_file_download_url = asset['browser_download_url']

            # 触发更新信号，即使程序版本是最新但有数据模板更新
            if program_download_url or data_file_download_url:
                self.update_available.emit(latest_version, program_download_url, description, data_file_version or "未知", data_file_download_url)
            self.update_check_finished.emit()
        except requests.RequestException as e:
            self.update_check_failed.emit(str(e))

    def is_newer_version(self, latest, current):
        # 简单的字符串比较，可能需要根据版本号格式进行更复杂的比较
        return latest > current


class UpdateDialog(QDialog):
    def __init__(self, parent, version_info, program_download_url, description):
        super().__init__(parent)
        self.setWindowTitle("发现新版本")
        layout = QVBoxLayout()

        # 使用传入的version_info显示版本信息
        layout.addWidget(QLabel(version_info))
        
        # 多行文本框显示版本描述，保持不变
        description_edit = QTextEdit()
        description_edit.setText(description)
        description_edit.setReadOnly(True)
        layout.addWidget(description_edit)

        # 更新按钮
        update_button = QPushButton("更新")
        update_button.clicked.connect(lambda: self.accept())
        layout.addWidget(update_button)

        # 取消按钮
        cancel_button = QPushButton("取消")
        cancel_button.clicked.connect(self.reject)
        layout.addWidget(cancel_button)

        self.setLayout(layout)
        self.resize(400, 300)




class MainWindow(QWidget):
    shift_times_updated = pyqtSignal(QTime, QTime, QTime, QTime)
    def __init__(self, version, data_file_version):
        super().__init__()

        # 版本信息
        self.version_info_path = "version_info.json"
        self.current_version = version
        self.data_file_version = data_file_version  # 处理的数据文件版本
        self.update_date = datetime.now().strftime("%Y-%m-%d")
        self.executable_name = os.path.basename(sys.argv[0])
        self.generate_version_file()

        # 设置窗口标题，包括应用名称、当前版本号和适用的数据模板版本号
        self.setWindowTitle(f"OMM测量数据自动文件转换 v{self.current_version} - 适用模板 v{self.data_file_version}")

        self.template_path = ""
        self.folder_path = ""
        self.judgement_type = ""  # 初始化 judgement_type
        self.selected_method = "最大值"  # 初始化 selected_method 为默认值
        self.range_limit = ""
        self.machine_number = ""  # 初始化机台号属性
        self.file_hashes = {}  # 用于追踪文本文件的哈希值
        self.consider_xy_center = True  # 这个是跟踪“XY中心坐标”开关状态的属性
        self.is_running = False
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.check_and_convert)
        self.toasts = []  # 存储当前显示的 Toasts
        self.conversion_thread = ConversionThread(self.template_path, self.folder_path, self.file_hashes, self.judgement_type, self.selected_method, self.range_limit, self.machine_number, self.consider_xy_center)

        self.shift_times_updated.connect(self.conversion_thread.update_shift_times)

        self.conversion_thread.file_converted.connect(self.file_converted_slot)  # 连接信号

        self.folder_list_dialog = None  # 用于跟踪文件夹列表窗口
        self.initUI()

        self.load_paths_from_file()  # 加载上次的路径

        self.last_keypress_times = []  # 存储最近按键的时间戳
        self.debug_sequence = []  # 存储最近按键序列

        # 加载设置
        self.load_settings()

        xy_center_state = self.load_xy_center_state()
        self.xy_center_coordinate_checkbox.setChecked(xy_center_state)
        self.xy_center_coordinate_checkbox.stateChanged.connect(self.xy_center_state_changed)

        # 为方法选择框添加事件监听器
        self.method_combo.currentTextChanged.connect(self.save_settings)

        # 为类型输入框添加事件监听器
        self.type_input.textChanged.connect(self.save_settings)

        # 为范围输入框添加事件监听器
        self.range_input1.textChanged.connect(self.save_settings)
        self.range_input2.textChanged.connect(self.save_settings)
        self.range_input3.textChanged.connect(self.save_settings)
        self.range_input4.textChanged.connect(self.save_settings)
        self.range_input5.textChanged.connect(self.save_settings)
        self.range_input6.textChanged.connect(self.save_settings)

        # 为机台输入框添加文本变化的事件监听器
        self.machine_input.textChanged.connect(self.update_machine_number)

        self.check_for_updates()


    def find_current_data_file_version(self):
        """在当前目录下查找所有数据文件并提取最新的版本号"""
        pattern = re.compile(r'数据模板(\d+\.\d+)|v(\d+\.\d+)\.xlsx', re.IGNORECASE)
        latest_version = "未知"
        
        for filename in os.listdir('.'):
            match = pattern.search(filename)
            if match:
                # 尝试从两个不同的捕获组获取版本号
                version = match.group(1) or match.group(2)
                if latest_version == "未知" or self.is_newer_version(version, latest_version):
                    latest_version = version
        return latest_version

    def is_newer_version(self, version1, version2):
        """比较两个版本号字符串，确定version1是否比version2新"""
        return tuple(map(int, version1.split('.'))) > tuple(map(int, version2.split('.')))


    def generate_version_file(self):
        """生成包含版本信息的文件"""
        version_info = {
            "version": self.current_version,
            "data_file_version": self.data_file_version,  # 添加数据文件版本信息
            "update_date": self.update_date,
            "executable_name": self.executable_name
        }
        with open(self.version_info_path, "w") as file:
            json.dump(version_info, file, indent=4)

    def check_for_updates(self):
        # GitHub的API路径，获取最新发布信息
        self.update_checker = UpdateCheckerThread(
            "https://api.github.com/repos/LAYccc03/OMM_excel/releases/latest",
            self.current_version)  # 传递当前版本号
        self.update_checker.update_available.connect(self.prompt_for_update)
        self.update_checker.update_check_finished.connect(self.on_update_check_finished)
        self.update_checker.update_check_failed.connect(self.on_update_check_failed)
        self.update_checker.start()
        self.show_update_checking_message()



    def show_update_checking_message(self):
        self.checking_update_msg = QMessageBox(self)
        self.checking_update_msg.setWindowTitle("检查更新")
        self.checking_update_msg.setText("正在检查更新，请稍候...")
        self.checking_update_msg.setStandardButtons(QMessageBox.NoButton)
        self.checking_update_msg.show()

    def on_update_check_finished(self):
        self.checking_update_msg.accept()

    def on_update_check_failed(self, error_message):
        # 创建QMessageBox实例而不是直接调用静态方法
        msgBox = QMessageBox(self)
        msgBox.setIcon(QMessageBox.Critical)
        msgBox.setWindowTitle("更新检查失败")
        msgBox.setText(error_message)
        msgBox.setStandardButtons(QMessageBox.Ok)  # 设置一个OK按钮
        msgBox.show()

        self.checking_update_msg.accept()  # 关闭正在检查更新的对话框
    
        # 使用QTimer来延迟关闭消息框
        QTimer.singleShot(2000, msgBox.accept)  # 2000毫秒后关闭


    def prompt_for_update(self, latest_version, program_download_url, description, latest_data_file_version, data_file_download_url):
        current_data_file_version_in_program = self.find_current_data_file_version()  # 获取当前程序位置下的数据文件版本

        version_info = (f"当前程序版本：{self.current_version}\n"
                        f"当前版本推荐数据模板：{self.data_file_version}\n"
                        f"当前程序内识别到的数据模板版本：{current_data_file_version_in_program}\n\n"
                        f"发现新程序版本：{latest_version}\n"
                        f"新版本推荐数据模板：{latest_data_file_version}")

        # 检查程序是否已是最新版本
        if self.current_version == latest_version and data_file_download_url:
            # 程序版本最新，检查数据模板是否需要更新
            self.prompt_for_data_template_update(latest_data_file_version, data_file_download_url)
        elif self.current_version != latest_version:
            # 程序有新版本，处理程序更新逻辑
            dialog = UpdateDialog(self, version_info, program_download_url, description)
            accepted = dialog.exec_() == QDialog.Accepted
            if accepted:
                self.download_and_update(program_download_url, data_file_download_url)
            else:
                # 用户取消程序更新，但可能需要更新数据模板
                self.prompt_for_data_template_update(latest_data_file_version, data_file_download_url)





    def download_and_update(self, program_download_url=None, data_file_download_url=None):
        try:
            program_file_path = None
            data_template_file_path = None

            if program_download_url:
                # 下载新版本程序
                program_file_path = self.download_file(program_download_url, "程序")
        
            if data_file_download_url:
                # 检查并下载数据模板
                data_template_file_path = self.download_file(data_file_download_url, "数据模板")

            if program_file_path:
                # 如果有程序文件下载，替换旧版本的程序并启动新版本
                # 此时，如果也有数据模板下载，则一并处理
                self.replace_and_restart(program_file_path, os.path.basename(program_file_path), data_template_file_path)
            elif data_template_file_path:
                # 如果只有数据模板文件下载，直接移动到程序所在目录
                self.move_file_to_program_dir(data_template_file_path)

        except requests.RequestException as e:
            QMessageBox.critical(self, "下载失败", f"下载新版本失败：{e}")


    def prompt_for_data_template_update(self, latest_data_file_version, data_file_download_url):
        current_data_file_version = self.find_current_data_file_version()
        if current_data_file_version == "未知" or self.is_newer_version(latest_data_file_version, current_data_file_version):
            reply = QMessageBox.question(self, "数据模板更新可用",
                                         f"当前数据模板版本：{current_data_file_version}\n\n"
                                         f"发现新数据模板版本：{latest_data_file_version}\n\n"
                                         "是否现在下载新数据模板？",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.Yes:
                self.download_and_update(None, data_file_download_url)  # 只更新数据模板


    def move_file_to_program_dir(self, file_path):
        """将文件移动到程序所在目录"""
        destination_path = os.path.join(os.path.dirname(os.path.abspath(self.executable_name)), os.path.basename(file_path))
        shutil.move(file_path, destination_path)
        QMessageBox.information(self, "更新完成", f"数据模板已下载最新版本并保存在程序目录：{destination_path}")


    def download_file(self, download_url, file_type):
        """通用文件下载方法"""
        try:
            response = requests.get(download_url, stream=True)
            response.raise_for_status()

            filename = download_url.split('/')[-1]
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, filename)

            total_length = int(response.headers.get('content-length', 0))
            progress_dialog = QProgressDialog(f"正在下载新{file_type}...", "取消", 0, total_length, self)
            progress_dialog.setWindowTitle(f"下载{file_type}更新")
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setAutoClose(True)
            progress_dialog.show()

            downloaded = 0
            start_time = time.time()
            with open(file_path, "wb") as file:
                for data in response.iter_content(chunk_size=4096):
                    downloaded += len(data)
                    file.write(data)
                    elapsed_time = time.time() - start_time
                    download_speed = downloaded / elapsed_time / 1024 / 1024  # MB/s
                    progress_dialog.setValue(downloaded)
                    progress_dialog.setLabelText(f"下载中: {downloaded/1024/1024:.2f}MB / {total_length/1024/1024:.2f}MB    下载速度：{download_speed:.2f}MB/s")
                    QApplication.processEvents()
            
                    if progress_dialog.wasCanceled():
                        QMessageBox.warning(self, f"下载{file_type}取消", f"{file_type}更新下载已取消。")
                        return None

            progress_dialog.close()
            return file_path  # 返回下载文件的路径
        except requests.RequestException as e:
            QMessageBox.critical(self, "下载失败", f"下载{file_type}失败：{e}")
            return None




    def replace_and_restart(self, new_executable_path, new_filename, new_data_template_path=None):
        current_dir = os.path.dirname(os.path.abspath(self.executable_name))
        new_executable_final_path = os.path.join(current_dir, new_filename)

        data_template_commands = ""
        if new_data_template_path:
            new_data_template_final_path = os.path.join(current_dir, os.path.basename(new_data_template_path))
            # 打印路径以供调试
            data_template_commands = f"""
            echo 数据模板源路径: {new_data_template_path}
            echo 数据模板目标路径: {new_data_template_final_path}
            echo 更新数据模板...
            del /f /q "{new_data_template_final_path}"
            move /y "{new_data_template_path}" "{new_data_template_final_path}"
            """
        print(data_template_commands)
        script_content = f"""
        @echo off
        echo 更新程序...
        ping 127.0.0.1 -n 6 > nul
        del /f /q "{os.path.join(current_dir, self.executable_name)}"
        move /y "{new_executable_path}" "{new_executable_final_path}"
        {data_template_commands}
        start "" "{new_executable_final_path}"
        rd /s /q "{os.path.dirname(new_executable_path)}"  # 删除临时文件夹及其内容
        del "%~f0"  # 删除本批处理脚本
        """

        script_path = os.path.join(tempfile.gettempdir(), "update_script.bat")
        with open(script_path, "w") as script_file:
            script_file.write(script_content)

        subprocess.Popen(["cmd.exe", "/c", script_path], cwd=current_dir, close_fds=True)
        self.close()









    def save_shift_times_to_file(self):
        times = {
            'day_shift_start': self.day_shift_start_time_edit.time().toString(),
            'day_shift_end': self.day_shift_end_time_edit.time().toString(),
            'night_shift_start': self.night_shift_start_time_edit.time().toString(),
            'night_shift_end': self.night_shift_end_time_edit.time().toString()
        }
        with open('shift_times.json', 'w') as file:
            json.dump(times, file)
#1--------------------------------------------------------UI--------------------------------------------------------1#
    def initUI(self):
        self.resize(800, 480)  # 设置默认大小
        layout = QVBoxLayout()
        layout.setSpacing(10)  # 调整控件之间的间距，根据需要可以减少这个值

        # 添加用于显示文件数量的标签和机台输入框的水平布局
        topLayout = QHBoxLayout()

        # 用于显示文件数量的标签
        self.file_count_label = QLabel('已转换数量：0/0', self)
        font = QFont('Microsoft YaHei', 10)
        self.file_count_label.setFont(font)
        topLayout.addWidget(self.file_count_label, alignment=Qt.AlignLeft)

        # 添加弹性空间以使机台标签和输入框靠右
        topLayout.addStretch()

        # 机台标签
        machine_label = QLabel('检验机台:', self)
        machine_label.setFont(font)
        topLayout.addWidget(machine_label, alignment=Qt.AlignRight)

        # 机台输入框
        self.machine_input = QLineEdit(self)
        self.machine_input.setFixedSize(QSize(100, 30))  # 设置输入框的大小
        topLayout.addWidget(self.machine_input, alignment=Qt.AlignRight)
        self.machine_input.textChanged.connect(self.update_machine_settings_json)  # 连接信号到槽函数

        # 将包含“已转换数量”标签和机台输入框的水平布局添加到主布局
        layout.addLayout(topLayout)

        # 读取机台设置
        self.load_machine_settings()

        # 创建一个水平布局用于图标按钮
        icons_layout = QHBoxLayout()

        # 第一个图标按钮
        search_icon = QPushButton()
        icon_path = resource_path('search_icon.png')
        search_icon.setIcon(QIcon(icon_path))
        search_icon.setIconSize(QSize(30, 30))
        search_icon.setFixedSize(QSize(50, 50))
        search_icon.clicked.connect(self.open_folder_list_dialog)
        icons_layout.addWidget(search_icon)

        # 第二个图标按钮
        second_icon = QPushButton()
        second_icon_path = resource_path('reset_icon.png')
        second_icon.setIcon(QIcon(second_icon_path))
        second_icon.setIconSize(QSize(30, 30))
        second_icon.setFixedSize(QSize(50, 50))
        second_icon.clicked.connect(self.reset_and_restart)
        icons_layout.addWidget(second_icon)

        # 添加弹性空间以使按钮靠左
        icons_layout.addStretch(1)

        # 将图标按钮的布局添加到主布局
        layout.addLayout(icons_layout)

        # 设置运行按钮
        self.run_button = QPushButton('RUN', self)
        self.run_button.clicked.connect(self.run_clicked)
        self.run_button.setFixedSize(QSize(220, 110))
        font = QFont('Microsoft YaHei', 18, QFont.Bold)
        self.run_button.setFont(font)

        layout.addStretch(1)
        # 先添加运行按钮
        layout.addWidget(self.run_button, 0, Qt.AlignCenter)

        # 添加一个具有负边距的垂直空间调节器来“推”色块向上
        spacer = QSpacerItem(20, -5, QSizePolicy.Minimum, QSizePolicy.Fixed)
        layout.addItem(spacer)

        # 创建色块
        self.colorBlock = ColorBlockWidget(self)
        self.colorBlock.setFixedSize(QSize(218, 7))  # 色块的宽度与按钮相等，高度为5
        layout.addWidget(self.colorBlock, 0, Qt.AlignCenter)

        layout.addStretch(1)

        # “恢复默认时间”按钮和提示文字的布局
        reset_layout = QHBoxLayout()
        reset_layout.addStretch(1)
        info_label = QLabel('范围外默认"白班"', self)
        info_label.setStyleSheet("color: #808080;")
        reset_layout.addWidget(info_label)
        self.reset_time_button = QPushButton('恢复成默认时间', self)
        self.reset_time_button.clicked.connect(self.reset_default_times)
        reset_layout.addWidget(self.reset_time_button)

        # 使用变量来初始化时间编辑器
        day_shift_start, day_shift_end = self.get_shift_time('day_shift_start', QTime(8, 0)), self.get_shift_time('day_shift_end', QTime(17, 30))
        night_shift_start, night_shift_end = self.get_shift_time('night_shift_start', QTime(20, 0)), self.get_shift_time('night_shift_end', QTime(8, 0))

        # 创建一个水平布局用于“方法”和“范围外默认'白班'”
        method_layout = QHBoxLayout()
        method_label = QLabel('方法:', self)
        self.method_combo = QComboBox(self)
        self.method_combo.addItems(['最大值', '最小值'])
        self.method_combo.setFixedWidth(100)
        method_layout.addWidget(method_label)
        method_layout.addWidget(self.method_combo)
        method_layout.addStretch(1)
        method_layout.addLayout(reset_layout)
        layout.addLayout(method_layout)

        # 创建一个水平布局用于“类型”和“白班时间范围”
        type_layout = QHBoxLayout()
        type_label = QLabel('判断类型:', self)
        self.type_input = QLineEdit(self)  # 将输入框设置为类属性以便在其他函数中访问
        self.type_input.setPlaceholderText("请输入类型")
        self.type_input.setFixedWidth(100)
        type_layout.addWidget(type_label)
        type_layout.addWidget(self.type_input)

        # 添加“XY中心坐标”开关
        self.xy_center_coordinate_checkbox = QCheckBox("XY中心坐标", self)
        self.xy_center_coordinate_checkbox.stateChanged.connect(self.xy_center_state_changed)
        type_layout.addWidget(self.xy_center_coordinate_checkbox)

        type_layout.addStretch(1)
        self.day_shift_start_time_edit, self.day_shift_end_time_edit = self.add_time_edit_layout('白班时间范围:', day_shift_start, day_shift_end, type_layout)
        layout.addLayout(type_layout)

        # 创建一个水平布局用于“限制范围”和“夜班时间范围”
        range_layout = QHBoxLayout()

        # 添加“限制范围”标签
        range_label = QLabel('限制范围:', self)
        range_layout.addWidget(range_label)

        # 创建并添加三个输入框
        self.range_input1 = QLineEdit(self)
        self.range_input1.setPlaceholderText("范围1")
        self.range_input1.setFixedWidth(60)  # 设置为60像素宽
        self.range_input1.setValidator(CustomValidator())  # 设置为浮点数验证器

        self.range_input2 = QLineEdit(self)
        self.range_input2.setPlaceholderText("范围2")
        self.range_input2.setFixedWidth(60)  # 设置为60像素宽
        self.range_input2.setValidator(CustomValidator())  # 设置为浮点数验证器

        self.range_input3 = QLineEdit(self)
        self.range_input3.setPlaceholderText("范围3")
        self.range_input3.setFixedWidth(60)  # 设置为60像素宽
        self.range_input3.setValidator(CustomValidator())  # 设置为浮点数验证器

        self.range_input4 = QLineEdit(self)
        self.range_input4.setPlaceholderText("范围4")
        self.range_input4.setFixedWidth(60)
        self.range_input4.setValidator(CustomValidator())

        self.range_input5 = QLineEdit(self)
        self.range_input5.setPlaceholderText("范围5")
        self.range_input5.setFixedWidth(60)
        self.range_input5.setValidator(CustomValidator())

        self.range_input6 = QLineEdit(self)
        self.range_input6.setPlaceholderText("范围6")
        self.range_input6.setFixedWidth(60)
        self.range_input6.setValidator(CustomValidator())


        # 将输入框添加到布局中
        range_layout.addWidget(self.range_input1)
        range_layout.addWidget(self.range_input2)
        range_layout.addWidget(self.range_input3)
        range_layout.addWidget(self.range_input4)
        range_layout.addWidget(self.range_input5)
        range_layout.addWidget(self.range_input6)


        # 添加弹性空间和夜班时间范围的布局
        range_layout.addStretch(1)
        self.night_shift_start_time_edit, self.night_shift_end_time_edit = self.add_time_edit_layout('夜班时间范围:', night_shift_start, night_shift_end, range_layout)
        layout.addLayout(range_layout)

        # 模板路径输入框和按钮
        self.template_line_edit = QLineEdit(self)
        self.template_line_edit.setPlaceholderText("选择模板路径")
        self.template_button = QPushButton('选择模板文件', self)
        self.template_button.clicked.connect(self.browse_template)
        layout.addWidget(self.template_line_edit)
        layout.addWidget(self.template_button)

        # 文件夹路径输入框和按钮
        self.folder_line_edit = QLineEdit(self)
        self.folder_line_edit.setPlaceholderText("选择需转换的文件夹路径")
        self.folder_button = QPushButton('选择自动转换文件夹', self)
        self.folder_button.clicked.connect(self.browse_folder)
        layout.addWidget(self.folder_line_edit)
        layout.addWidget(self.folder_button)

        self.setLayout(layout)
#2--------------------------------------------------------UI--------------------------------------------------------2#
    # 定义点击事件处理函数
    def open_folder_list_dialog(self):
        # 检查路径是否有效
        if not self.folder_path or not os.path.exists(self.folder_path):
            QMessageBox.warning(self, "路径无效", "请选择有效的路径。")
            return

        if self.folder_list_dialog is not None:
            self.folder_list_dialog.raise_()
        else:
            self.folder_list_dialog = FolderListDialog(self.folder_path)
            self.folder_list_dialog.setAttribute(Qt.WA_DeleteOnClose)

            # 设置FolderListDialog的位置在MainWindow的中间
            main_window_geom = self.frameGeometry()
            dialog_width = self.folder_list_dialog.width()
            dialog_height = self.folder_list_dialog.height()

            # 计算使FolderListDialog居中的位置
            dialog_x = main_window_geom.left() + (main_window_geom.width() - dialog_width) / 2
            dialog_y = main_window_geom.top() + (main_window_geom.height() - dialog_height) / 2

            self.folder_list_dialog.move(int(dialog_x), int(dialog_y))

            self.folder_list_dialog.show()
            self.folder_list_dialog.finished.connect(self.on_folder_list_dialog_closed)




    def on_folder_list_dialog_closed(self):
        self.folder_list_dialog = None  # 释放对话框实例，以便下次创建新实例

    def reset_default_times(self):
        self.day_shift_start_time_edit.setTime(QTime(8, 0))
        self.day_shift_end_time_edit.setTime(QTime(17, 30))
        self.night_shift_start_time_edit.setTime(QTime(20, 0))
        self.night_shift_end_time_edit.setTime(QTime(8, 0))
        self.save_shift_times_to_file()  # 可选：将默认时间保存到文件

    def get_shift_time(self, key, default_time):
        try:
            with open('shift_times.json', 'r') as file:
                times = json.load(file)
                time_str = times[key]

                time = QTime.fromString(time_str, "HH:mm:ss")  # 匹配包含秒的格式
                if not time.isValid():
                    return default_time
                return time
        except Exception as e:
            print(f"Error loading time for {key}: {e}")  # 打印错误信息
            return default_time


    def add_time_edit_layout(self, label_text, start_time, end_time, layout):
        # 添加时间编辑器和它们的标签到布局中
        label = QLabel(label_text, self)
        start_time_edit = QTimeEdit(self)
        start_time_edit.setDisplayFormat("HH:mm")
        start_time_edit.setTime(start_time)
        end_time_edit = QTimeEdit(self)
        end_time_edit.setDisplayFormat("HH:mm")
        end_time_edit.setTime(end_time)

        # 设置时间编辑器宽度为窗口宽度的1/6
        start_time_edit.setMaximumWidth(self.width() // 6)
        end_time_edit.setMaximumWidth(self.width() // 6)

        # 添加到布局中
        layout.addWidget(label)
        layout.addWidget(start_time_edit)
        layout.addWidget(end_time_edit)

        start_time_edit.timeChanged.connect(self.update_shift_times)
        end_time_edit.timeChanged.connect(self.update_shift_times)

        return start_time_edit, end_time_edit

    def update_shift_times(self):
        day_start = self.day_shift_start_time_edit.time()
        day_end = self.day_shift_end_time_edit.time()
        night_start = self.night_shift_start_time_edit.time()
        night_end = self.night_shift_end_time_edit.time()
        self.shift_times_updated.emit(day_start, day_end, night_start, night_end)

        self.save_shift_times_to_file()

#3--------------------------------------------------------Debug--------------------------------------------------------3#
    def keyPressEvent(self, event):
        # 检测按键，并添加到序列中
        current_time = time.time()
        self.debug_sequence.append(event.text())
        self.last_keypress_times.append(current_time)

        # 检查是否在5秒内依次输入了 "debug"
        if ''.join(self.debug_sequence[-5:]).lower() == "debug" and \
           current_time - self.last_keypress_times[-5] < 5:
            self.extract_dimensions()
            self.debug_sequence.clear()  # 重置序列
            self.last_keypress_times.clear()  # 重置时间戳

        super().keyPressEvent(event)  # 调用基类的键盘事件处理

    def extract_dimensions(self):
        dimensions = {}  # 字典，用于存储 Dimension 名称和文件名集合
        for file_name in os.listdir(self.folder_path):
            if file_name.lower().endswith('.xlsx'):
                try:
                    wb = load_workbook(os.path.join(self.folder_path, file_name), read_only=True)
                    ws = wb.active
                    rows = find_rows(ws)
                    if rows:
                        row_dimension = rows[0]
                        print(f"维度所在行: {row_dimension}")  # 调试打印
                        col_start = 4
                        while True:
                            cell = ws.cell(row=row_dimension, column=col_start)
                            if not cell.value:
                                break
                            if isinstance(cell.value, str) and ':' in cell.value:
                                dimension_value = cell.value.split(':', 1)[1].strip()
                                if dimension_value not in dimensions:
                                    dimensions[dimension_value] = set()
                                dimensions[dimension_value].add(file_name)
                            col_start += 1
                    else:
                        print(f"未找到维度行: {file_name}")  # 如果没有找到维度行
                    wb.close()
                    print(f"处理完成: {file_name}")  # 打印处理完成的文件名
                except Exception as e:
                    print(f"处理文件 {file_name} 时发生错误: {e}")

        if dimensions:
            # 使用 DimensionsDialog 显示结果
            dialog = DimensionsDialog(dimensions, self.folder_path)
            dialog.exec_()
        else:
            print("未找到任何维度信息。")
#4--------------------------------------------------------Debug--------------------------------------------------------4#

    def save_paths_to_file(self):
            paths = {
                "template_path": self.template_path,
                "folder_path": self.folder_path
            }
            with open('paths.json', 'w') as file:
                json.dump(paths, file)

    def load_paths_from_file(self):
        try:
            with open('paths.json', 'r') as file:
                paths = json.load(file)
                self.template_path = paths.get('template_path', '')
                self.folder_path = paths.get('folder_path', '')
                self.template_line_edit.setText(self.template_path)
                self.folder_line_edit.setText(self.folder_path)
        except FileNotFoundError:
            pass  # 文件未找到时不做任何操作

    def file_converted_slot(self, file_name):
        # 被转换线程在转换每个文件后调用以更新文件数量
        self.update_file_count(len(self.file_hashes))

    def update_file_count(self, converted_count=None):
        # 检查文件夹是否存在
        if not os.path.exists(self.folder_path):
            print(f"文件夹 {self.folder_path} 不存在。")
            self.file_count_label.setText('已转换数量：0/0')  # 更新显示为0
            return

        # 获取当前文件夹的文件总数量
        if self.folder_path:
            total_count = len([name for name in os.listdir(self.folder_path) if name.lower().endswith('.txt')])
            if converted_count is None:
                converted_count = 0  # 初始状态下已转换的文件数量为0
            self.file_count_label.setText(f'已转换数量：{converted_count}/{total_count}')



    def browse_template(self):
        self.template_path = QFileDialog.getOpenFileName(self, '选择模板文件', '', 'Excel files (*.xlsx)')[0]
        if self.template_path:  # 确保用户选中了文件
            self.template_line_edit.setText(self.template_path)
            self.save_paths_to_file()  # 保存路径

    def reset_conversion_data(self):
        # 重置已转换文件的列表和文件计数
        self.file_hashes.clear()
        self.update_file_count(0)
        # 重置转换线程中的文件夹路径
        self.conversion_thread.folder_path = self.folder_path


    def closeEvent(self, event):
        # 当 MainWindow 关闭时，也关闭 FolderListDialog（如果它是打开的）
        if self.folder_list_dialog is not None:
            self.folder_list_dialog.close()

        super().closeEvent(event)  # 调用父类的 closeEvent 方法


    def reset_and_restart(self):
        # 清除和重置数据
        self.reset_conversion_data()

        # 如果正在运行，则停止并重新开始
        if self.is_running:
            self.timer.stop()
            self.timer.start(3000)
        
    def browse_folder(self):
        selected_folder = QFileDialog.getExistingDirectory(self, '选择需转换的txt文件夹')
        if selected_folder and selected_folder != self.folder_path:
            self.folder_path = selected_folder
            self.folder_line_edit.setText(self.folder_path)
            self.reset_conversion_data()
            self.save_paths_to_file()  # 保存路径
            # 更新文件夹列表对话框的路径
            if self.folder_list_dialog:
                self.folder_list_dialog.set_folder_path(self.folder_path)
                self.folder_list_dialog.load_folder_list()
            if self.is_running:
                self.timer.stop()
                self.timer.start(3000)

    def run_clicked(self):
        if not self.is_running:
            if self.template_path and self.folder_path:
                self.is_running = True
                self.run_button.setText("STOP")
                self.colorBlock.setGradientColors('#f5af19', '#f12711', is_breathing=False)  # 设置为滚动渐变
                self.timer.start(3000)
                self.update_file_count()  # 获取并显示文件总数量
            else:
                self.show_toast("请先选择文件路径")
        else:
            self.is_running = False
            self.run_button.setText("RUN")
            self.colorBlock.setGradientColors('#36D1DC', '#5B86E5')  # 恢复呼吸渐变
            self.timer.stop()

    def show_toast(self, message):
        toast = Toast(message, width=200, height=50, parent=self)
    
        # 设置投影效果
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(25)
        shadow.setXOffset(5)
        shadow.setYOffset(5)

        # 设置投影的颜色和透明度（alpha 值设置为 128，半透明）
        shadow.setColor(QColor(0, 0, 0, 50))
    
        toast.setGraphicsEffect(shadow)

        self.toasts.append(toast)
        self.update_toast_positions()
        toast.show_toast()
        QTimer.singleShot(4000, lambda toast=toast: self.remove_toast(toast) if toast in self.toasts else None)


    def load_xy_center_state(self):
        try:
            with open('ui_settings.json', 'r') as f:
                settings = json.load(f)
                return settings.get('xy_center_coordinate', False)  # 默认为False
        except FileNotFoundError:
            return False  # 如果文件不存在，也默认为False


    def save_xy_center_state(self):
        settings = {'xy_center_coordinate': self.xy_center_coordinate_checkbox.isChecked()}
        with open('ui_settings.json', 'w') as f:
            json.dump(settings, f)

    # 假设这个方法用于处理“XY中心坐标”复选框状态的改变
    def xy_center_state_changed(self, state):
        self.consider_xy_center = state == Qt.Checked
        self.save_xy_center_state()

    def check_and_convert(self):
        # 更新 judgement_type、selected_method、range_limit 和 machine_number 的值
        self.update_judgement_type()
        self.update_method()
        self.update_range()
        self.update_machine_number()

        # 获取“XY中心坐标”复选框的状态
        consider_xy_center = self.xy_center_coordinate_checkbox.isChecked()

        day_start = self.day_shift_start_time_edit.time()
        day_end = self.day_shift_end_time_edit.time()
        night_start = self.night_shift_start_time_edit.time()
        night_end = self.night_shift_end_time_edit.time()

        # 创建 ConversionThread，同时传递XY中心坐标的考虑状态
        self.conversion_thread = ConversionThread(self.template_path, self.folder_path, self.file_hashes, self.judgement_type, self.selected_method, self.range_limit, self.machine_number, consider_xy_center)
        self.conversion_thread.update_shift_times(day_start, day_end, night_start, night_end)
        self.conversion_thread.conversion_done.connect(self.on_conversion_done)
        self.conversion_thread.start()


    def remove_toast(self, toast):
        # 检查 Toast 是否还存在于 toasts 列表中
        if toast in self.toasts:
            # 如果存在，则从列表中移除并删除 Toast 对象
            self.toasts.remove(toast)
            toast.deleteLater()
            self.update_toast_positions()
        # 如果 Toast 已经不在列表中，那么它已经被删除，不需要再次进行删除


    def load_machine_settings(self):
        settings_file = 'machine_settings.json'
        default_machine_number = '1'
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r') as file:
                    settings = json.load(file)
                    machine_number = settings.get('machine', default_machine_number)
            except json.JSONDecodeError:
                # 如果json文件损坏，设置默认机台号
                machine_number = default_machine_number
        else:   
            machine_number = default_machine_number
        
        # 将读取的机台号或默认值设置到输入框
        self.machine_input.setText(machine_number)

    def update_machine_number(self):
        self.machine_number = self.machine_input.text()
        self.update_machine_settings_json()  # 同时更新和保存机台设置到JSON文件

    def update_machine_settings_json(self):
        machine_text = self.machine_input.text().strip()  # 获取机台输入框的内容，去除两端空白
        # 如果输入框为空，则设置默认机台号为1
        machine_text = machine_text if machine_text else '1'
        machine_settings = {'machine': machine_text}  # 创建或更新机台设置字典
        with open('machine_settings.json', 'w') as file:  # 打开或创建 JSON 文件并写入数据
            json.dump(machine_settings, file, indent=4)



    def update_range(self):
        # 获取三个输入框的文本内容
        range1 = self.range_input1.text()
        range2 = self.range_input2.text()
        range3 = self.range_input3.text()
        range4 = self.range_input4.text()
        range5 = self.range_input5.text()
        range6 = self.range_input6.text()

        # 更新合并字符串，包括新的输入框
        self.range_limit = "/".join(filter(None, [range1, range2, range3, range4, range5, range6]))

    def save_settings(self):
        # 获取三个输入框的文本内容
        range1 = self.range_input1.text()
        range2 = self.range_input2.text()
        range3 = self.range_input3.text()
        range4 = self.range_input4.text()
        range5 = self.range_input5.text()
        range6 = self.range_input6.text()

        # 更新combined_range字符串
        combined_range = "/".join(filter(None, [range1, range2, range3, range4, range5, range6]))

        settings = {
            "method": self.method_combo.currentText(),
            "type": self.type_input.text(),
            "range": combined_range
        }

        with open("settings.json", "w") as file:
            json.dump(settings, file)


    def load_settings(self):
        default_settings = {
            "method": "最大值",
            "type": "直径/偏移量",
            "range": ["<0.5", "<0.525", "=1.175", "=1.275", "=1.375", ">1.85"]  # 添加默认值为空的额外范围
        }

        try:
            with open("settings.json", "r") as file:
                settings = json.load(file)

                # 获取方法设置，如果为空则使用默认值
                self.method_combo.setCurrentText(settings.get("method", default_settings["method"]))

                # 获取类型设置，如果为空则使用默认值
                self.type_input.setText(settings.get("type", default_settings["type"]))

                # 获取范围设置，如果为空则使用默认值
                range_settings = settings.get("range", default_settings["range"])
                # 确保range_settings是列表格式
                if isinstance(range_settings, str):
                    range_settings = range_settings.split('/')
                self.range_input1.setText(range_settings[0] if len(range_settings) > 0 else "")
                self.range_input2.setText(range_settings[1] if len(range_settings) > 1 else "")
                self.range_input3.setText(range_settings[2] if len(range_settings) > 2 else "")
                self.range_input4.setText(range_settings[3] if len(range_settings) > 3 else "")
                self.range_input5.setText(range_settings[4] if len(range_settings) > 4 else "")
                self.range_input6.setText(range_settings[5] if len(range_settings) > 5 else "")

        except FileNotFoundError:
            # 设置文件未找到时使用默认值
            self.method_combo.setCurrentText(default_settings["method"])
            self.type_input.setText(default_settings["type"])
            self.range_input1.setText(default_settings["range"][0])
            self.range_input2.setText(default_settings["range"][1])
            self.range_input3.setText(default_settings["range"][2])
            self.range_input4.setText(default_settings["range"][3])
            self.range_input5.setText(default_settings["range"][4])
            self.range_input6.setText(default_settings["range"][5])



    def update_method(self):
        # 获取当前在下拉菜单中选中的值
        self.selected_method = self.method_combo.currentText()

    def update_judgement_type(self):
        """更新 judgement_type 的值"""
        self.judgement_type = self.type_input.text()


    def on_conversion_done(self):
        # 转换全部完成时调用
        self.update_file_count(len(self.file_hashes))  # 更新文件数量显示
        message = "转换完成"
        # 检查是否已经存在相同的消息
        if any(toast.text() == message for toast in self.toasts):
            return  # 如果已经有相同的 Toast 显示，则不再添加新的

        toast = Toast(message, width=200, height=60, parent=self)
        # 设置投影效果
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(25)
        shadow.setXOffset(5)
        shadow.setYOffset(5)
        shadow.setColor(QColor(0, 0, 0, 50))  # 设置投影的颜色和透明度
        toast.setGraphicsEffect(shadow)

        self.toasts.append(toast)
        self.update_toast_positions()
        toast.show_toast()
        QTimer.singleShot(2000, lambda toast=toast: self.remove_toast(toast) if toast in self.toasts else None)


    def update_toast_positions(self):
        x = self.width() - 200 - 5  # 假设每个 Toast 的宽度为 200
        y = 45  # 初始位置
        for existing_toast in self.toasts:
            existing_toast.move(x, y)
            y += existing_toast.height() + 2  # 上一个 Toast 的高度 + 5px 间隔

    def remove_toast(self, toast):
        if toast in self.toasts:
            self.toasts.remove(toast)
            if toast:  # 这里检查 toast 对象是否仍然有效
                toast.deleteLater()
        self.update_toast_positions()



class DetailsDialog(QDialog):
    def __init__(self, folder_path, parent=None):
        super().__init__(parent, Qt.FramelessWindowHint)  # 添加无边框窗口标志
        self.folder_path = folder_path
        self.initUI()

    def initUI(self):
        self.setWindowTitle("图片预览")
        self.resize(200, self.parent().height())

        # 设置窗口背景透明
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setStyleSheet("background:transparent;")

        # 初始化定时器
        self.slideshow_timer = QTimer(self)
        self.slideshow_timer.timeout.connect(self.next_image)
        self.slideshow_timer.start(1500)  # 设置为2秒切换

        # 首先初始化主布局
        self.layout = QVBoxLayout(self)

        # 创建并添加图片标签
        self.image_label = QLabel(self)
        self.image_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.image_label)


    def leaveEvent(self, event):
        # 鼠标离开子窗口时关闭窗口
        self.close()
        super().leaveEvent(event)

    def update_details(self, folder_name):
        self.image_paths = []  # 存储图片路径的列表
        folder_path = os.path.join(self.folder_path, folder_name)
        for file_name in os.listdir(folder_path):
            if file_name.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                self.image_paths.append(os.path.join(folder_path, file_name))

        self.current_image_index = 0  # 设置当前图片索引
        if self.image_paths:
            self.show_image(QPixmap(self.image_paths[self.current_image_index]))
            return True
        return False

    def next_image(self):
        # 轮播到下一张图片
        if self.image_paths:
            self.current_image_index = (self.current_image_index + 1) % len(self.image_paths)
            self.show_image(QPixmap(self.image_paths[self.current_image_index]))

            # 在这里调用更新位置的方法
            self.update_position()

    def update_position(self):
        # 这里应该包含更新 DetailsDialog 位置的代码
        if hasattr(self.parent(), 'update_details_dialog_position'):
            self.parent().update_details_dialog_position()

    def show_image(self, pixmap):
        # 显示图片，高度填满整个窗口，等比例缩放，宽度自适应
        scaled_pixmap = pixmap.scaledToHeight(self.height(), Qt.SmoothTransformation)
        self.image_label.setPixmap(scaled_pixmap)

        # 调整窗口宽度以适应图片
        self.setFixedWidth(scaled_pixmap.width())
        self.adjustSize()  # 调整窗口大小以适应图片

        # 设置图片标签的样式，添加#F0F0F0颜色的边框和5像素的圆角
        self.image_label.setStyleSheet("""
            border: 4px solid #F0F0F0;
            border-radius: 5px;
            background-color: #1A212D;
        """)

    def reset_slideshow_timer(self):
        # 重置轮播定时器
        self.slideshow_timer.start(1500)


class FolderListDialog(QDialog):
    def __init__(self, folder_path, parent=None):
        super().__init__(parent)
        self.folder_path = folder_path
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.load_folder_list)
        self.outside_check_timer = QTimer(self)
        self.outside_check_timer.timeout.connect(self.check_mouse_position)
        self.outside_check_timer.start(1000)  # 每1000毫秒检查一次
        self.rename_dialog = None  # 初始化重命名对话框引用

        # 添加文件夹状态监测定时器
        self.folder_monitor_timer = QTimer(self)
        self.folder_monitor_timer.timeout.connect(self.check_folder_changes)
        self.folder_monitor_timer.start(1500)  # 每 1.5 秒检测一次

        # 记录初始文件夹状态
        self.initial_folder_info = self.record_all_folder_info()

        self.initUI()

        # 启用拖放
        self.setAcceptDrops(True)

    def set_folder_path(self, new_folder_path):
        self.folder_path = new_folder_path
        if hasattr(self, 'details_dialog'):
            self.details_dialog.folder_path = new_folder_path

    def check_mouse_position(self):
        # 检查鼠标是否在父窗口和子窗口范围外
        if hasattr(self, 'details_dialog'):
            if not (self.rect().contains(self.mapFromGlobal(QCursor.pos())) or
                    self.details_dialog.rect().contains(self.details_dialog.mapFromGlobal(QCursor.pos()))):
                self.details_dialog.close()

    def mousePressEvent(self, event):
        # 点击父窗口空白区域时关闭子窗口
        if hasattr(self, 'details_dialog'):
            self.details_dialog.close()
        super().mousePressEvent(event)



    def initUI(self):
        self.setWindowTitle("文件快速检索工具")
        self.resize(640, 380)

        layout = QVBoxLayout(self)

        # 添加置顶开关复选框
        self.always_on_top_checkbox = QCheckBox("保持窗口置顶", self)
        self.always_on_top_checkbox.stateChanged.connect(self.toggle_always_on_top)
        layout.addWidget(self.always_on_top_checkbox)

        # 添加搜索栏
        self.search_line_edit = QLineEdit(self)
        self.search_line_edit.setPlaceholderText("搜索文件夹")
        self.search_line_edit.textChanged.connect(self.search_folder)
        layout.addWidget(self.search_line_edit)

        # 添加列表控件
        self.list_widget = QListWidget(self)
        layout.addWidget(self.list_widget)

        # 双击信号连接到新的槽函数
        self.list_widget.itemDoubleClicked.connect(self.copy_name_to_rename_dialog)

        self.list_widget.itemClicked.connect(self.show_details_dialog)

        # 为 list_widget 添加鼠标右键事件
        self.list_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.list_widget.customContextMenuRequested.connect(self.show_context_menu)

        self.load_folder_list()

    def copy_name_to_rename_dialog(self, item):
        print("双击事件触发")  # 调试信息
        full_folder_name = item.text().split(' | ')[-1]  # 提取完整的文件夹名称

        # 如果是Excel文件，去除后缀
        folder_name = os.path.splitext(full_folder_name)[0] if full_folder_name.lower().endswith('.xlsx') else full_folder_name

        if self.rename_dialog and not self.rename_dialog.isHidden():
            print("重命名窗口已打开且未隐藏")  # 调试信息
            line_edit = self.rename_dialog.findChild(QLineEdit)  # 找到重命名窗口中的 QLineEdit
            if line_edit:
                line_edit.setText(folder_name)
            else:
                print("未找到输入框")  # 调试信息
        else:
            print("重命名窗口未打开或已隐藏，将文件夹名称复制到剪贴板")  # 调试信息
            clipboard = QApplication.clipboard()
            clipboard.setText(folder_name)


    def search_folder(self):
        search_term = self.search_line_edit.text().lower()
        if search_term:
            # 根据搜索关键词更新文件夹列表
            self.update_folder_list(search_term)
        else:
            # 搜索栏为空时恢复默认排序
            self.load_folder_list()



    def show_context_menu(self, position):
        # 获取鼠标点击的列表项
        list_item = self.list_widget.itemAt(position)
        if list_item:
            folder_name = list_item.text().split(' | ')[-1]

            # 创建上下文菜单
            context_menu = QMenu(self)
            open_excel_action = QAction("打开 Excel", self)
            open_folder_action = QAction("打开文件夹", self)
            context_menu.addAction(open_excel_action)
            context_menu.addAction(open_folder_action)

            # 连接动作信号到槽函数
            open_excel_action.triggered.connect(lambda: self.open_similar_excel(folder_name))
            open_folder_action.triggered.connect(lambda: self.open_folder(folder_name))

            # 显示菜单
            context_menu.exec_(self.list_widget.mapToGlobal(position))

    def open_folder(self, folder_name):
        folder_path = os.path.join(self.folder_path, folder_name)
        if os.path.exists(folder_path):
            # 在不同的操作系统中打开文件夹
            if os.name == 'nt':  # Windows
                os.startfile(folder_path)
            elif os.name == 'posix':  # macOS, Linux
                subprocess.Popen(['open' if sys.platform == 'darwin' else 'xdg-open', folder_path])

    def open_similar_excel(self, folder_name):
        # 寻找与文件夹名称最相似的 Excel 文件
        excel_path, similarity = self.find_most_similar_excel(folder_name)
        if similarity >= 0.7 and os.path.exists(excel_path):
            if os.name == 'nt':
                os.startfile(excel_path)
            elif os.name == 'posix':
                subprocess.Popen(['open' if sys.platform == 'darwin' else 'xdg-open', excel_path])
        else:
            QMessageBox.information(self, "信息", "未找到相似的 Excel 文件或文件不存在。")

    def find_most_similar_excel(self, folder_name):
        highest_similarity = 0.0
        most_similar_path = ""
        for file_name in os.listdir(self.folder_path):
            if file_name.lower().endswith('.xlsx'):
                similarity = difflib.SequenceMatcher(None, folder_name, file_name).ratio()
                if similarity > highest_similarity:
                    highest_similarity = similarity
                    most_similar_path = os.path.join(self.folder_path, file_name)
        return most_similar_path, highest_similarity




    def show_details_dialog(self, item):
        folder_name = item.text().split(' | ')[-1]  # 获取文件夹名称
        folder_path = os.path.join(self.folder_path, folder_name)

        # 检查路径是否为目录
        if not os.path.isdir(folder_path):
            return

        if not hasattr(self, 'details_dialog'):
            self.details_dialog = DetailsDialog(self.folder_path, self)
        else:
            # 确保更新了 details_dialog 的文件夹路径
            self.details_dialog.folder_path = self.folder_path

        image_found = self.details_dialog.update_details(folder_name)
        if image_found:
            self.details_dialog.show()  # 如果找到图片，则显示子窗口
            self.update_details_dialog_position()
            # 重置轮播定时器
            self.details_dialog.reset_slideshow_timer()
        else:
            self.details_dialog.close()  # 如果没有找到图片，则关闭子窗口


    def moveEvent(self, event):
        super().moveEvent(event)
        self.update_details_dialog_position()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_details_dialog_position()

    def update_details_dialog_position(self):
        # 更新子窗口位置
        if hasattr(self, 'details_dialog'):
            parent_frame = self.frameGeometry()

            # 获取并保持子窗口的当前宽度
            child_width = self.details_dialog.width()
            new_x = parent_frame.left() - child_width
            new_y = parent_frame.top()

            # 设置子窗口的高度为父窗口的2/3
            child_height = int(parent_frame.height() * 2 / 3)
            self.details_dialog.setFixedHeight(child_height)  # 只固定高度
            self.details_dialog.move(new_x, new_y)


    def count_files(self, folder_path, extensions):
        count = 0
        for file in os.listdir(folder_path):
            if os.path.splitext(file)[1].lower() in extensions:
                count += 1
        return count

    def toggle_always_on_top(self, state):
        if state == Qt.Checked:
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        else:
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowStaysOnTopHint)
        self.show()  # 重新显示窗口以应用新的窗口标志

    def load_folder_list(self):
        # 获取当前文件夹状态
        current_folder_info = self.record_all_folder_info()

        self.list_widget.clear()  # 清空当前列表
        folder_info_list = []  # 存储文件夹和图片数量
        excel_info_list = []  # 单独存储Excel文件信息

        if os.path.isdir(self.folder_path):
            for item in os.listdir(self.folder_path):
                # 跳过以 '~$' 开头的临时文件
                if item.startswith('~$'):
                    continue

                full_path = os.path.join(self.folder_path, item)
                if os.path.isdir(full_path):
                    image_count = self.count_files(full_path, ['.jpg', '.jpeg', '.png', '.gif'])
                    other_count = self.count_files(full_path, ['.txt', '.pdf', '.docx'])
                    folder_info_list.append((item, image_count, other_count))
                elif item.lower().endswith('.xlsx'):
                    excel_info_list.append(item)

        # 按图片数量排序文件夹，然后按名称排序Excel文件
        folder_info_list.sort(key=lambda x: x[1])
        excel_info_list.sort()

        # 过滤掉与文件夹名称高度相似的Excel文件
        filtered_excel_info_list = []
        for excel_file in excel_info_list:
            excel_file_name = os.path.splitext(excel_file)[0]
            if not any(difflib.SequenceMatcher(None, excel_file_name, folder[0]).ratio() >= 0.8 for folder in folder_info_list):
                filtered_excel_info_list.append(excel_file)

        # 将文件夹信息添加到列表控件
        for item, image_count, other_count in folder_info_list:
            list_item_text = f"图 {image_count}"
            if other_count > 0:
                list_item_text += f" | 其它 {other_count}"
            list_item_text += f" | {item}"

            list_item = QListWidgetItem(list_item_text)
            if item in self.initial_folder_info and self.initial_folder_info[item] != current_folder_info[item]:
                list_item.setBackground(QColor('#FFBA5D'))  # 使用自定义高亮颜色

            self.list_widget.addItem(list_item)

        # 将过滤后的Excel文件信息添加到列表控件
        for excel_file in filtered_excel_info_list:
            list_item = QListWidgetItem(f"Excel | {excel_file}")
            self.list_widget.addItem(list_item)

        # 更新初始文件夹状态
        self.initial_folder_info = current_folder_info



    def update_folder_list(self, search_term):
        self.list_widget.clear()  # 清空当前列表
        folder_info_list = []  # 存储文件夹名称和图片数量
        excel_info_list = []  # 单独存储Excel文件信息

        if os.path.isdir(self.folder_path):
            for item in os.listdir(self.folder_path):
                # 跳过以 '~$' 开头的临时文件
                if item.startswith('~$'):
                    continue

                full_path = os.path.join(self.folder_path, item)
                if os.path.isdir(full_path):
                    image_count = self.count_files(full_path, ['.jpg', '.jpeg', '.png', '.gif'])
                    other_count = self.count_files(full_path, ['.txt', '.pdf', '.docx'])
                    folder_info_list.append((item, image_count, other_count))
                elif item.lower().endswith('.xlsx'):
                    excel_info_list.append(item)

        # 过滤和排序文件夹列表
        filtered_folder_list = [info for info in folder_info_list if search_term in info[0].lower()]
        filtered_folder_list.sort(key=lambda x: x[0].lower().find(search_term))

        # 过滤掉与文件夹名称高度相似的Excel文件
        filtered_excel_info_list = []
        for excel_file in excel_info_list:
            excel_file_name = os.path.splitext(excel_file)[0]
            if not any(difflib.SequenceMatcher(None, excel_file_name, folder[0]).ratio() >= 0.8 for folder in filtered_folder_list):
                if search_term in excel_file.lower():
                    filtered_excel_info_list.append(excel_file)
        filtered_excel_info_list.sort()

        # 将过滤后的文件夹信息添加到列表控件
        for item, image_count, other_count in filtered_folder_list:
            list_item = f"图 {image_count}"
            if other_count > 0:
                list_item += f" | 其它 {other_count}"
            list_item += f" | {item}"
            self.list_widget.addItem(list_item)

        # 将过滤后的Excel文件信息添加到列表控件
        for excel_file in filtered_excel_info_list:
            list_item = QListWidgetItem(f"Excel | {excel_file}")
            self.list_widget.addItem(list_item)



    def check_folder_changes(self):
        """检查文件夹状态是否有变化，并更新列表"""
        current_folder_info = self.record_all_folder_info()
        if current_folder_info != self.initial_folder_info:
            print("检测到文件夹状态变化，更新列表。")
            self.load_folder_list()
            self.initial_folder_info = self.record_all_folder_info()  # 重新记录初始状态

    def record_all_folder_info(self):
        """记录所有文件夹的初始状态"""
        folder_info = {}

        # 检查路径是否为空或不存在
        if not self.folder_path or not os.path.exists(self.folder_path):
            print("无效的路径或路径不存在。")
            return folder_info

        for folder_name in os.listdir(self.folder_path):
            folder_path = os.path.join(self.folder_path, folder_name)
            if os.path.isdir(folder_path):
                folder_info[folder_name] = self.record_folder_info(folder_path)

        return folder_info


    def record_folder_info(self, folder_path):
        """记录单个文件夹的文件数量和哈希值"""
        file_count = 0
        file_hashes = {}
        for file_name in os.listdir(folder_path):
            if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                file_count += 1
                file_path = os.path.join(folder_path, file_name)
                file_hash = self.get_file_hash(file_path)
                file_hashes[file_name] = file_hash
        return {'file_count': file_count, 'file_hashes': file_hashes}

    def get_file_hash(self, file_path):
        """计算文件的哈希值"""
        hash_md5 = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    # 快捷键粘贴图片
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_V and event.modifiers() == Qt.ControlModifier:
            print("快捷键粘贴事件捕获")
            self.paste_image_from_clipboard()
        else:
            super().keyPressEvent(event)

    def paste_image_from_clipboard(self):
        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()

        if mime_data.hasImage():
            print("剪贴板中有图片")
            image = clipboard.image()
            self.prompt_and_save_image(image, "png")  # 假设剪贴板中的图像是 PNG 格式
        elif mime_data.hasUrls():
            urls = mime_data.urls()
            if urls and urls[0].isLocalFile():
                file_path = urls[0].toLocalFile()
                self.handle_dropped_file(file_path)
        else:
            print("剪贴板中没有图片或文件路径")

    # 拖放图片
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            file_path = urls[0].toLocalFile()
            self.handle_dropped_file(file_path)

    def handle_dropped_file(self, file_path):
        # 提取原文件名的扩展名
        extension = os.path.splitext(file_path)[1].lstrip('.')
        # 调用重命名对话框，不传递默认名称
        self.prompt_and_save_image(file_path, extension)

    def prompt_and_save_image(self, image_path, extension):
        # 创建对话框和布局
        dialog = QDialog(self)
        dialog.setModal(False)  # 设置为非模态
        layout = QVBoxLayout(dialog)

        # 添加用于预览图片的 QLabel
        image_label = QLabel(dialog)
        original_name = ""
        if isinstance(image_path, QImage):
            pixmap = QPixmap.fromImage(image_path)
        else:
            pixmap = QPixmap(image_path)
            original_name = os.path.splitext(os.path.basename(image_path))[0]

        image_label.setPixmap(pixmap.scaled(380, 380, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        image_label.setAlignment(Qt.AlignCenter)  # 图片水平居中
        layout.addWidget(image_label)

        # 输入框和标签
        input_label = QLabel("输入图片新名称（留空默认使用原文件名）:", dialog)
        layout.addWidget(input_label)
        line_edit = QLineEdit(dialog)
        layout.addWidget(line_edit)

        # 添加确定和取消按钮
        buttons = QDialogButtonBox(Qt.Horizontal, dialog)
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")
        buttons.addButton(ok_button, QDialogButtonBox.AcceptRole)
        buttons.addButton(cancel_button, QDialogButtonBox.RejectRole)
        layout.addWidget(buttons)

        # 更新重命名对话框的引用
        self.rename_dialog = dialog

        # 连接对话框的关闭信号
        dialog.finished.connect(lambda: setattr(self, 'rename_dialog', None))

        # 连接按钮信号
        ok_button.clicked.connect(lambda: save_image(line_edit.text()))
        cancel_button.clicked.connect(lambda: cancel_and_close())

        def save_image(new_name):
            if not new_name:
                if original_name:
                    new_name = original_name
                else:
                    new_name = QDateTime.currentDateTime().toString("yyyyMMddHHmmss")

            # 寻找与图片名相似的文件夹
            most_similar_folder = self.find_most_similar_folder(new_name)
            base_path = self.folder_path if not most_similar_folder else os.path.join(self.folder_path, most_similar_folder)
            new_image_path = os.path.join(base_path, new_name + '.' + extension)

            # 检查同名文件是否存在，并添加序号
            counter = 1
            while os.path.exists(new_image_path):
                new_image_path = os.path.join(base_path, f"{new_name}_{counter}.{extension}")
                counter += 1

            pixmap.save(new_image_path)
            self.load_folder_list()
            self.rename_dialog = None  # 初始化重命名对话框引用
            dialog.deleteLater()  # 点击确定后删除对话框

        def cancel_and_close():
            self.rename_dialog = None
            dialog.deleteLater()

        # 显示对话框
        dialog.setLayout(layout)
        dialog.setWindowTitle("重命名图片")
        dialog.show()


    def find_most_similar_folder(self, image_name, similarity_threshold=0.6):
        most_similar = None
        highest_similarity = 0.0
        for folder in os.listdir(self.folder_path):
            folder_path = os.path.join(self.folder_path, folder)
            if os.path.isdir(folder_path):
                similarity = difflib.SequenceMatcher(None, folder, image_name).ratio()
                if similarity > highest_similarity and similarity > similarity_threshold:
                    most_similar = folder
                    highest_similarity = similarity
        return most_similar


#1--------------------------------------------------------Debug--------------------------------------------------------1#
class DimensionsDialog(QDialog):
    def __init__(self, dimensions, folder_path, parent=None):
        super(DimensionsDialog, self).__init__(parent)
        self.dimensions = dimensions
        self.folder_path = folder_path  # 添加 folder_path 属性
        self.setStyleSheet("background-color: white;")
        self.initUI()

    def initUI(self):
        layout = QGridLayout(self)
        max_items_per_column = 5
        col = 0
        row = 0

        for idx, (dimension, files) in enumerate(self.dimensions.items(), 1):
            # 如果当前列的项目数已经达到上限，则换到下一列
            if row >= max_items_per_column:
                col += 1
                row = 0

            # 添加带序列号的 Dimension 名称标签
            dimension_label = QLabel(f'{idx}. {dimension}', self)
            dimension_label.mousePressEvent = lambda event, dim=dimension: self.showDimensionFiles(dim)
            layout.addWidget(dimension_label, row, col)
            row += 1

        self.setLayout(layout)

    def showDimensionFiles(self, dimension):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"文件包含 {dimension}")
        dialog.setStyleSheet("background-color: white;")
        layout = QGridLayout(dialog)
        max_items_per_column = 5
        col = 0
        row = 0

        files = self.dimensions[dimension]
        for idx, file_name in enumerate(files, 1):
            file_path = os.path.join(self.folder_path, file_name)
            file_label = QLabel(f'{idx}. {file_name}', dialog)
            file_label.mousePressEvent = lambda event, path=file_path: self.openFile(path)
            layout.addWidget(file_label, row, col)
            row += 1

            spacer_label = QLabel("", dialog)
            layout.addWidget(spacer_label, row, col)
            row += 1

        dialog.setLayout(layout)
        dialog.exec_()

    def openFile(self, file_path):
        if os.path.exists(file_path):
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS, Linux
                    subprocess.run(['open', file_path], check=True)
            except Exception as e:
                QMessageBox.warning(self, "打开文件错误", str(e))
        else:
            QMessageBox.warning(self, "错误", f"文件 {file_path} 不存在")
#2--------------------------------------------------------Debug--------------------------------------------------------2#


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow(current_version, data_file_version)
    window.show()
    sys.exit(app.exec_())